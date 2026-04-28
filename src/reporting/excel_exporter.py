"""
Excel 报告导出器。

使用 openpyxl 将结构化数据写入多 Sheet 的 .xlsx 文件，
支持样式美化和图表嵌入。
"""

from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from loguru import logger
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.reporting.data_extractor import ExtractedData
from src.reporting.report_templates import COLLECTOR_LABELS, get_report_template, is_structured_template


# ==================== 样式常量 ====================

HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_BORDER = Border(
    bottom=Side(style="thin", color="1F3864"),
    right=Side(style="thin", color="D9E2F3"),
)

DATA_FONT = Font(name="微软雅黑", size=10)
DATA_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
DATA_BORDER = Border(
    bottom=Side(style="hair", color="D9E2F3"),
    right=Side(style="hair", color="D9E2F3"),
)

ALT_ROW_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")


def export_to_excel(
    data: ExtractedData,
    output_path: str | Path,
    title: str = "GamedataAutoFlux 数据报告",
    llm_content: str | None = None,
    template_id: str = "default",
    template_validation: dict[str, Any] | None = None,
) -> Path:
    """
    将提取后的数据导出为 Excel 文件。

    Args:
        data: DataExtractor 提取的结构化数据
        output_path: 输出文件路径
        title: 报告标题（写入 Sheet 标题行）
        llm_content: LLM 生成的分析文本（可选）

    Returns:
        生成的 Excel 文件路径
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if is_structured_template(template_id):
        return _export_structured_template(
            data=data,
            output_path=output_path,
            title=title,
            template_id=template_id,
            template_validation=template_validation or {},
            llm_content=llm_content,
        )

    wb = Workbook()

    # 删除默认 Sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Sheet 1: 游戏概览
    if data.overview:
        _write_overview_sheet(wb, data.overview, title)

    # Sheet 2: 评论明细
    if data.reviews:
        _write_reviews_sheet(wb, data.reviews)

    # Sheet 3: 趋势数据
    if template_id in {"general_game", "steam_game"}:
        _write_unified_trends_sheet(wb, data)
    elif data.trends:
        _write_trends_sheet(wb, data.trends)

    # Sheet 4: 相关搜索词
    if data.related_queries:
        _write_related_queries_sheet(wb, data.related_queries)

    # Sheet 5: LLM 分析
    if llm_content:
        _write_llm_sheet(wb, llm_content, title)

    # 如果没有任何数据，创建空的说明页
    if not wb.sheetnames:
        ws = wb.create_sheet("说明")
        ws["A1"] = "暂无可用数据。请先执行采集任务并落库后再生成报告。"
        ws["A1"].font = Font(name="微软雅黑", size=12, bold=True)

    wb.save(str(output_path))
    logger.info(f"[ExcelExporter] 报告已生成: {output_path} ({len(wb.sheetnames)} sheets)")
    return output_path


# ==================== 模板化工作簿 ====================

def _export_structured_template(
    data: ExtractedData,
    output_path: Path,
    title: str,
    template_id: str,
    template_validation: dict[str, Any],
    llm_content: str | None,
) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    template = get_report_template(template_id)
    if template_id == "general_game":
        _write_operations_monitor_sheet(wb, data)

    _write_template_summary_sheet(wb, title, template_id, template_validation)

    if llm_content:
        _write_llm_sheet(wb, llm_content, title)

    if data.overview:
        _write_table_sheet(wb, "核心概览", data.overview)

    if data.steam_player_peaks and template_id in {"general_game", "steam_game"}:
        _write_steam_peak_sheet(wb, data.steam_player_peaks, title)

    if data.steam_monthly_peaks and template_id in {"general_game", "steam_game"}:
        _write_line_series_sheet(
            wb,
            "Steam月峰值",
            data.steam_monthly_peaks,
            date_header="月份",
            value_header="Peak在线人数",
            chart_title="SteamDB 月峰值（1Y）",
            y_axis_title="Peak Players",
        )

    if data.google_trends and template_id in {"general_game", "steam_game"}:
        _write_line_series_sheet(
            wb,
            "Google趋势",
            data.google_trends,
            date_header="日期",
            value_header="热度值",
            chart_title="Google Trends 搜索热度趋势",
            y_axis_title="Trend Value",
        )

    if data.monitor_metrics and template_id in {"general_game", "steam_game"}:
        _write_monitor_sheet(wb, data.monitor_metrics)

    if data.events and template_id in {"general_game", "steam_game"}:
        _write_table_sheet(wb, "事件数据", data.events, max_width=80)

    if data.community_discussions and template_id in {"general_game", "steam_game"}:
        _write_table_sheet(wb, "社区讨论", data.community_discussions, max_width=80)

    taptap_reviews = [row for row in data.reviews if row.get("数据来源") == "TapTap"]
    other_reviews = [row for row in data.reviews if row.get("数据来源") != "TapTap"]
    if taptap_reviews:
        _write_table_sheet(wb, "TapTap评论", taptap_reviews, max_width=70)
    if other_reviews:
        _write_table_sheet(wb, "评论与讨论", other_reviews, max_width=70)

    if data.related_queries:
        _write_related_queries_sheet(wb, data.related_queries)

    if template_id in {"general_game", "steam_game"}:
        _write_unified_trends_sheet(wb, data)
    elif data.trends:
        _write_trends_sheet(wb, data.trends)

    _write_raw_appendices(wb, data.raw_sources)

    if template and template_validation.get("status") == "partial":
        ws = wb["报告摘要"]
        ws["D1"] = "缺失数据源提示"
        ws["D1"].font = HEADER_FONT
        ws["D1"].fill = HEADER_FILL
        ws["D2"] = (
            "本报告允许在数据源不完整时生成。缺失部分会影响结论置信度，"
            "请以附录中的原始 JSON 为准。"
        )
        ws["D2"].alignment = DATA_ALIGNMENT
        ws.column_dimensions["D"].width = 52

    wb.save(str(output_path))
    logger.info(f"[ExcelExporter] 模板报告已生成: {output_path} ({len(wb.sheetnames)} sheets)")
    return output_path


def _write_template_summary_sheet(
    wb: Workbook,
    title: str,
    template_id: str,
    template_validation: dict[str, Any],
) -> None:
    template = get_report_template(template_id)
    rows = [
        {"项目": "报告标题", "内容": title},
        {"项目": "生成时间", "内容": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {
            "项目": "模板",
            "内容": f"{template.name} ({template.id})" if template else template_id,
        },
        {
            "项目": "模板状态",
            "内容": "完整" if template_validation.get("status") == "complete" else "部分数据",
        },
        {
            "项目": "必需数据源",
            "内容": _format_collectors(template_validation.get("required_collectors", [])),
        },
        {
            "项目": "已选数据源",
            "内容": _format_collectors(template_validation.get("available_collectors", [])),
        },
        {
            "项目": "缺失数据源",
            "内容": _format_collectors(template_validation.get("missing_collectors", [])) or "无",
        },
    ]
    _write_table_sheet(wb, "报告摘要", rows, max_width=80)


def _write_operations_monitor_sheet(wb: Workbook, data: ExtractedData) -> None:
    """Write the A 在运营产品监测 sheet based on the reference brief template."""
    ws = wb.create_sheet("A 在运营产品监测", 0)
    _setup_operations_monitor_layout(ws)

    game_name = _first_non_empty(row.get("游戏名") for row in data.overview) or _first_non_empty(
        row.get("游戏名") for row in data.events
    )
    steam_row = _find_overview_row(data.overview, "Steam")
    qimai_row = _find_overview_row(data.overview, "Qimai")
    monitor_row = _find_overview_row(data.overview, "Monitor")
    gtrends_row = _find_overview_row(data.overview, "Google Trends")

    pc_headers = [
        "产品",
        "整体好评率",
        "近期好评率(30 Days)",
        "3个月好评率趋势图",
        "7日ccu peak",
        "30日CCU peak",
        "3个月ccu趋势",
        "Google Trends（3个月趋势图）",
        "twitch tracker(7天平均观看人数)",
        "twitch tracker(90天趋势)",
        "steam畅销榜",
        " Wishlist Activity(7d Gain)/Follower(7d Gain)",
    ]
    mobile_headers = [
        "产品",
        "iOS畅销榜CN",
        "iOS畅销榜CN趋势图",
        "App Store 评分CN",
        "App Store 每日评价趋势图",
        "DAU（30日平均）",
        "DAU 趋势（90日）",
        "下载（30日平均）",
        "下载 趋势（90日）",
        "收入（30日平均）",
        "收入USD趋势（90日）",
    ]
    event_headers = [
        "产品",
        "事件日期",
        "时间节点",
        "事件类型",
        "事件摘要",
        "一句话概括",
        "来源平台",
        "来源链接",
        "原文摘录",
    ]

    _write_operations_title(ws, 1, "在营产品", "A1:M1", font_size=20)
    _write_operations_title(ws, 3, "本期在营竞品总览（PC）", "A3:C3")
    _write_operations_header(ws, 4, pc_headers)
    _write_operations_row(
        ws,
        5,
        [
            game_name,
            _first_non_empty([steam_row.get("整体好评率"), steam_row.get("好评率")]),
            steam_row.get("近期好评率(30 Days)", ""),
            "",
            steam_row.get("7日ccu peak", ""),
            steam_row.get("30日CCU peak", ""),
            "",
            "",
            monitor_row.get("twitch tracker(7天平均观看人数)", ""),
            "",
            steam_row.get("steam畅销榜", ""),
            steam_row.get("WishList Activity(7d Gain)/Follower(7d Gain)", ""),
        ],
        len(pc_headers),
    )
    _add_operations_trend_chart(wb, ws, "Steam好评率", _operations_steam_review_rows(data), "D5")
    _add_operations_trend_chart(wb, ws, "Steam CCU", _operations_steam_ccu_rows(data), "G5")
    _add_operations_trend_chart(wb, ws, "Google Trends", _operations_google_rows(data), "H5")
    _add_operations_trend_chart(wb, ws, "Twitch", _operations_twitch_rows(data), "J5")

    _write_operations_title(ws, 8, "本期在营竞品总览（移动端）", "A8:C8")
    _write_operations_header(ws, 9, mobile_headers)
    _write_operations_row(
        ws,
        10,
        [
            game_name,
            qimai_row.get("Qimai grossing rank CN", ""),
            "",
            qimai_row.get("Qimai AppStore rating CN", ""),
            "",
            qimai_row.get("Qimai DAU avg 30d", ""),
            "",
            qimai_row.get("Qimai downloads avg 30d", ""),
            "",
            qimai_row.get("Qimai revenue avg 30d", ""),
            "",
        ],
        len(mobile_headers),
    )
    _add_operations_trend_chart(wb, ws, "iOS畅销榜", _operations_qimai_rows(data, "iOS grossing rank"), "C10")
    _add_operations_trend_chart(wb, ws, "AppStore评价", _operations_qimai_rows(data, "AppStore reviews"), "E10")
    _add_operations_trend_chart(wb, ws, "DAU", _operations_qimai_rows(data, "DAU"), "G10")
    _add_operations_trend_chart(wb, ws, "下载", _operations_qimai_rows(data, "Downloads"), "I10")
    _add_operations_trend_chart(wb, ws, "收入", _operations_qimai_rows(data, "Revenue"), "K10")

    _write_operations_title(ws, 14, "产品事件数据信息", "A14:J14")
    _write_operations_title(ws, 15, game_name or "", "A15:J15", fill_color="EAF2F8", font_color="000000", font_size=10)
    _write_operations_header(ws, 16, event_headers)

    row_index = 18
    for event in data.events[:40]:
        _write_operations_row(
            ws,
            row_index,
            [
                event.get("游戏名", game_name),
                event.get("日期", ""),
                event.get("时间节点", ""),
                event.get("事件类型", ""),
                event.get("摘要", ""),
                event.get("一句话概括", event.get("标题", "")),
                event.get("来源", event.get("来源平台", "")),
                event.get("URL", event.get("来源链接", "")),
                event.get("原文摘录", ""),
            ],
            len(event_headers),
            align="left",
        )
        row_index += 1

    ws.freeze_panes = "A4"


def _setup_operations_monitor_layout(ws) -> None:
    widths = {
        "A": 18,
        "B": 16.625,
        "C": 27.25,
        "D": 31.375,
        "E": 33.625,
        "F": 17.625,
        "G": 29.625,
        "H": 22,
        "I": 13,
        "J": 22,
        "K": 22,
        "L": 19.875,
        "M": 9,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for row, height in {1: 58.5, 2: 24, 3: 24, 4: 49.5, 5: 132, 8: 24, 10: 132}.items():
        ws.row_dimensions[row].height = height


def _write_operations_title(
    ws,
    row: int,
    text: str,
    merge_range: str,
    *,
    fill_color: str = "1F4E78",
    font_color: str = "FFFFFF",
    font_size: int = 14,
) -> None:
    ws.merge_cells(merge_range)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    cell.font = Font(name="微软雅黑", size=font_size, bold=True, color=font_color)
    cell.alignment = Alignment(horizontal="center" if row == 1 else "left", vertical="center", wrap_text=True)


def _write_operations_header(ws, row: int, headers: list[str]) -> None:
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_index, value=header)
        cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        cell.font = Font(name="微软雅黑", size=10, bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = DATA_BORDER


def _write_operations_row(
    ws,
    row: int,
    values: list[Any],
    width: int,
    *,
    align: str = "center",
) -> None:
    for col_index in range(1, width + 1):
        value = values[col_index - 1] if col_index <= len(values) else ""
        cell = ws.cell(row=row, column=col_index, value=value)
        cell.font = DATA_FONT
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        cell.border = DATA_BORDER


def _find_overview_row(rows: list[dict[str, Any]], source_keyword: str) -> dict[str, Any]:
    source_keyword = source_keyword.lower()
    exact_matches = {
        "steam": {"steam"},
        "qimai": {"qimai", "qimai(appstore)"},
        "monitor": {"monitor"},
        "google trends": {"google trends"},
    }
    allowed = exact_matches.get(source_keyword)
    if allowed:
        for row in rows:
            source = str(row.get("数据来源", "")).lower().strip()
            source_base = source.split("(", 1)[0].strip()
            if source in allowed or source_base in allowed:
                return row
    for row in rows:
        source = str(row.get("数据来源", "")).lower()
        if source_keyword in source:
            return row
    return {}


def _trend_reference(value: Any, fallback_sheet: str) -> str:
    if value not in (None, ""):
        text = str(value)
        if fallback_sheet and "见《" not in text:
            return f"{text} / 见《{fallback_sheet}》"
        return text
    return f"见《{fallback_sheet}》" if fallback_sheet else ""


def _add_operations_trend_chart(
    wb: Workbook,
    target_ws,
    title: str,
    rows: list[tuple[Any, Any]],
    anchor: str,
) -> None:
    cleaned = [(date, _numeric_value(value)) for date, value in rows if date not in (None, "")]
    cleaned = [(date, value) for date, value in cleaned if value is not None]
    if len(cleaned) < 2:
        return

    data_ws = _ensure_operations_trend_data_sheet(wb)
    start_row = data_ws.max_row + 2 if data_ws.max_row > 1 or data_ws["A1"].value else 1
    data_ws.cell(row=start_row, column=1, value=title)
    data_ws.cell(row=start_row + 1, column=1, value="日期")
    data_ws.cell(row=start_row + 1, column=2, value="值")
    for row_index, (date_value, numeric_value) in enumerate(cleaned, start=start_row + 2):
        data_ws.cell(row=row_index, column=1, value=str(date_value))
        data_ws.cell(row=row_index, column=2, value=numeric_value)

    chart = LineChart()
    chart.title = title
    chart.style = 2
    chart.legend = None
    chart.width = 4.6
    chart.height = 2.8
    chart.y_axis.majorGridlines = None
    chart.x_axis.tickLblPos = "none"
    chart.y_axis.tickLblPos = "none"
    chart.x_axis.majorTickMark = "none"
    chart.y_axis.majorTickMark = "none"
    chart.graphical_properties = None
    data_ref = Reference(data_ws, min_col=2, min_row=start_row + 1, max_row=start_row + 1 + len(cleaned))
    cats_ref = Reference(data_ws, min_col=1, min_row=start_row + 2, max_row=start_row + 1 + len(cleaned))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    target_ws.add_chart(chart, anchor)


def _ensure_operations_trend_data_sheet(wb: Workbook):
    sheet_name = "_A趋势图数据"
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.sheet_state = "hidden"
    return ws


def _numeric_value(value: Any) -> float | int | None:
    if isinstance(value, bool) or value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip().replace(",", "").replace("%", "")
    try:
        number = float(text)
    except ValueError:
        return None
    return int(number) if number.is_integer() else number


def _operations_steam_review_rows(data: ExtractedData) -> list[tuple[Any, Any]]:
    rows = [
        (row.get("日期"), row.get("热度值"))
        for row in data.trends
        if row.get("类型") == "Steam好评率(90天)"
    ]
    return sorted(rows, key=lambda item: _series_sort_key(item[0]))


def _operations_steam_ccu_rows(data: ExtractedData) -> list[tuple[Any, Any]]:
    rows = [(row.get("日期"), row.get("在线峰值")) for row in data.steam_player_peaks]
    return sorted(rows, key=lambda item: _series_sort_key(item[0]))


def _operations_google_rows(data: ExtractedData) -> list[tuple[Any, Any]]:
    rows = [(row.get("日期"), row.get("热度值")) for row in data.google_trends]
    return sorted(rows, key=lambda item: _series_sort_key(item[0]))


def _operations_twitch_rows(data: ExtractedData) -> list[tuple[Any, Any]]:
    rows = [
        (row.get("日期"), row.get("Twitch平均观看"))
        for row in data.monitor_metrics
        if row.get("Twitch平均观看") not in (None, "")
    ]
    return sorted(rows, key=lambda item: _series_sort_key(item[0]))


def _operations_qimai_rows(data: ExtractedData, metric_name: str) -> list[tuple[Any, Any]]:
    rows = [
        (row.get("日期"), row.get("值"))
        for row in data.trends
        if row.get("数据源") == "Qimai(AppStore)" and row.get("指标") == metric_name
    ]
    return sorted(rows, key=lambda item: _series_sort_key(item[0]))


def _qimai_trend_count(data: ExtractedData, metric_name: str) -> str:
    count = sum(
        1
        for row in data.trends
        if row.get("数据源") == "Qimai(AppStore)" and row.get("指标") == metric_name
    )
    return f"{count} points" if count else ""


def _write_table_sheet(
    wb: Workbook,
    sheet_name: str,
    rows: list[dict[str, Any]],
    max_width: int = 50,
):
    ws = wb.create_sheet(_unique_sheet_title(wb, sheet_name))
    if not rows:
        ws["A1"] = "暂无数据"
        return ws, []

    headers = _collect_headers(rows)
    _write_header_row(ws, headers, row_num=1)

    for row_index, row in enumerate(rows, start=2):
        for col_index, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=row.get(header, ""))
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT
            cell.border = DATA_BORDER
            if row_index % 2 == 0:
                cell.fill = ALT_ROW_FILL

    _auto_column_width(ws, headers, max_width=max_width)
    ws.freeze_panes = "A2"
    return ws, headers


def _write_steam_peak_sheet(wb: Workbook, rows: list[dict[str, Any]], report_title: str) -> None:
    sorted_rows = sorted(rows, key=lambda row: str(row.get("日期", "")))
    ws = wb.create_sheet(_unique_sheet_title(wb, "Steam在线峰值"))
    if not sorted_rows:
        ws["A1"] = "暂无 Steam 在线峰值数据"
        return

    game_name = _first_non_empty(row.get("游戏名") for row in sorted_rows)
    app_id = _first_non_empty(row.get("App ID") for row in sorted_rows)
    data_source = _first_non_empty(row.get("数据源") for row in sorted_rows) or "SteamDB"
    time_slice = _first_non_empty(row.get("时间粒度") for row in sorted_rows)
    peak_values = [
        int(row["在线峰值"])
        for row in sorted_rows
        if isinstance(row.get("在线峰值"), (int, float))
    ]
    dates = [str(row.get("日期", "")) for row in sorted_rows if row.get("日期")]
    max_peak = max(peak_values) if peak_values else ""
    min_peak = min(peak_values) if peak_values else ""
    avg_peak = round(sum(peak_values) / len(peak_values)) if peak_values else ""
    max_date = next((row.get("日期") for row in sorted_rows if row.get("在线峰值") == max_peak), "")
    min_date = next((row.get("日期") for row in sorted_rows if row.get("在线峰值") == min_peak), "")

    ws["A1"] = f"{game_name or report_title} SteamDB 近30天 Peak 在线人数"
    ws["A1"].font = Font(name="微软雅黑", bold=True, size=14, color="1F3864")
    ws.merge_cells("A1:F1")

    summary_rows = [
        ("游戏", game_name, "App ID", app_id, "数据区间", f"{dates[0]} 至 {dates[-1]}" if dates else ""),
        ("最高Peak", max_peak, "最高日期", max_date, "平均Peak", avg_peak),
        ("最低Peak", min_peak, "最低日期", min_date, "记录数", len(sorted_rows)),
        ("数据来源", data_source, "数据切片", time_slice, "生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for row_index, values in enumerate(summary_rows, start=2):
        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.alignment = DATA_ALIGNMENT
            if col_index in {1, 3, 5}:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            else:
                cell.font = DATA_FONT
                if row_index % 2 == 0:
                    cell.fill = ALT_ROW_FILL

    headers = ["日期", "Peak在线人数", "时间戳(UTC)", "较前日变化", "较前日变化率", "是否最高"]
    header_row = 8
    _write_header_row(ws, headers, row_num=header_row)

    previous_value: int | float | None = None
    for row_index, row in enumerate(sorted_rows, start=header_row + 1):
        value = row.get("在线峰值")
        diff = value - previous_value if isinstance(value, (int, float)) and isinstance(previous_value, (int, float)) else ""
        rate = (
            diff / previous_value
            if isinstance(diff, (int, float)) and isinstance(previous_value, (int, float)) and previous_value
            else ""
        )
        values = [
            row.get("日期", ""),
            value,
            row.get("时间戳(UTC)", ""),
            diff,
            rate,
            "是" if value == max_peak else "",
        ]
        for col_index, cell_value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=cell_value)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT
            cell.border = DATA_BORDER
            if col_index == 5 and isinstance(cell_value, (int, float)):
                cell.number_format = "0.00%"
            if row_index % 2 == 0:
                cell.fill = ALT_ROW_FILL
        if isinstance(value, (int, float)):
            previous_value = value

    _auto_column_width(ws, headers, max_width=28)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 26
    ws.freeze_panes = "A9"

    if len(sorted_rows) >= 2:
        chart = LineChart()
        chart.title = "SteamDB Peak 在线人数趋势"
        chart.y_axis.title = "Peak Players"
        chart.x_axis.title = "日期"
        chart.style = 10
        chart.width = 28
        chart.height = 14
        data_ref = Reference(ws, min_col=2, min_row=header_row, max_row=header_row + len(sorted_rows))
        cats_ref = Reference(ws, min_col=1, min_row=header_row + 1, max_row=header_row + len(sorted_rows))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, "H2")


def _write_line_series_sheet(
    wb: Workbook,
    sheet_name: str,
    rows: list[dict[str, Any]],
    date_header: str,
    value_header: str,
    chart_title: str,
    y_axis_title: str,
) -> None:
    sorted_rows = sorted(rows, key=lambda row: _series_sort_key(row.get(date_header)))
    ws, headers = _write_table_sheet(wb, sheet_name, sorted_rows)
    if len(sorted_rows) < 2 or date_header not in headers or value_header not in headers:
        return

    date_col = headers.index(date_header) + 1
    value_col = headers.index(value_header) + 1

    chart = LineChart()
    chart.title = chart_title
    chart.y_axis.title = y_axis_title
    chart.x_axis.title = date_header
    chart.style = 10
    chart.width = 28
    chart.height = 14

    data_ref = Reference(ws, min_col=value_col, min_row=1, max_row=len(sorted_rows) + 1)
    cats_ref = Reference(ws, min_col=date_col, min_row=2, max_row=len(sorted_rows) + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, "H2")


def _write_monitor_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    sorted_rows = sorted(rows, key=lambda row: _series_sort_key(row.get("日期")))
    ws, headers = _write_table_sheet(wb, "Monitor指标", sorted_rows)
    if len(sorted_rows) < 2 or "日期" not in headers:
        return

    value_headers = [header for header in ("Twitch平均观看", "Twitch峰值观看") if header in headers]
    if not value_headers:
        return

    date_col = headers.index("日期") + 1
    chart = LineChart()
    chart.title = "Twitch观看趋势"
    chart.y_axis.title = "Viewers"
    chart.x_axis.title = "日期"
    chart.style = 10
    chart.width = 28
    chart.height = 14

    for header in value_headers:
        value_col = headers.index(header) + 1
        data_ref = Reference(ws, min_col=value_col, min_row=1, max_row=len(sorted_rows) + 1)
        chart.add_data(data_ref, titles_from_data=True)

    cats_ref = Reference(ws, min_col=date_col, min_row=2, max_row=len(sorted_rows) + 1)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, f"A{len(sorted_rows) + 4}")


def _write_raw_appendices(wb: Workbook, raw_sources: list[dict[str, Any]]) -> None:
    for index, source in enumerate(raw_sources, start=1):
        collector = source.get("collector", "unknown")
        label = COLLECTOR_LABELS.get(str(collector), str(collector))
        sheet_name = f"附录{index}_{label}"
        rows = [
            {"JSON路径": path, "值": value}
            for path, value in _flatten_json(source.get("data"), max_rows=5000)
        ]
        if not rows:
            rows = [{"JSON路径": "$", "值": ""}]
        _write_table_sheet(wb, sheet_name, rows, max_width=90)


def _flatten_json(value: Any, path: str = "$", max_rows: int = 5000):
    emitted = 0
    stack: list[tuple[str, Any]] = [(path, value)]
    while stack and emitted < max_rows:
        current_path, current_value = stack.pop()
        if isinstance(current_value, dict):
            for key, child in reversed(list(current_value.items())):
                stack.append((f"{current_path}.{key}", child))
        elif isinstance(current_value, list):
            for index, child in reversed(list(enumerate(current_value))):
                stack.append((f"{current_path}[{index}]", child))
        else:
            emitted += 1
            yield current_path, _cell_text(current_value)
    if stack:
        yield "$.__truncated__", f"原始 JSON 过大，仅写入前 {max_rows} 个叶子节点"


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        text = str(value)
    else:
        text = json.dumps(value, ensure_ascii=False, default=str)
    return text[:30000]


def _format_collectors(collectors: list[str] | tuple[str, ...]) -> str:
    return ", ".join(COLLECTOR_LABELS.get(str(item), str(item)) for item in collectors)


def _first_non_empty(values) -> Any:
    for value in values:
        if value not in (None, ""):
            return value
    return ""


def _unique_sheet_title(wb: Workbook, title: str) -> str:
    safe = "".join("_" if ch in '[]:*?/\\' else ch for ch in title).strip() or "Sheet"
    safe = safe[:31]
    if safe not in wb.sheetnames:
        return safe
    base = safe[:28]
    counter = 2
    while True:
        candidate = f"{base}_{counter}"[:31]
        if candidate not in wb.sheetnames:
            return candidate
        counter += 1


# ==================== Sheet 写入函数 ====================

def _write_overview_sheet(wb: Workbook, rows: list[dict[str, Any]], title: str) -> None:
    """写入游戏概览 Sheet。"""
    ws = wb.create_sheet("游戏概览")
    headers = _collect_headers(rows)

    # 写标题行
    _write_header_row(ws, headers, row_num=1)

    # 写数据行
    for i, row in enumerate(rows, start=2):
        for j, header in enumerate(headers, start=1):
            cell = ws.cell(row=i, column=j, value=row.get(header, ""))
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT
            cell.border = DATA_BORDER
            if i % 2 == 0:
                cell.fill = ALT_ROW_FILL

    _auto_column_width(ws, headers)
    ws.freeze_panes = "A2"

    # 如果有数值列，生成图表
    _add_overview_chart(ws, headers, rows)


def _write_reviews_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    """写入评论明细 Sheet。"""
    ws = wb.create_sheet("评论明细")
    headers = _collect_headers(rows)

    _write_header_row(ws, headers, row_num=1)

    for i, row in enumerate(rows, start=2):
        for j, header in enumerate(headers, start=1):
            cell = ws.cell(row=i, column=j, value=row.get(header, ""))
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT
            cell.border = DATA_BORDER
            if i % 2 == 0:
                cell.fill = ALT_ROW_FILL

    _auto_column_width(ws, headers, max_width=60)
    ws.freeze_panes = "A2"


def _write_unified_trends_sheet(wb: Workbook, data: ExtractedData) -> None:
    """写入结构化模板使用的统一趋势数据 Sheet。

    A 在运营产品监测页的多个指标都会指向本 Sheet，因此这里按数据源拆成多个表：
    Steam 好评率、Steam CCU、Google Trends、Twitch Tracker、Qimai/AppStore。
    """
    sections: list[tuple[str, list[dict[str, Any]], list[str], list[str]]] = []

    steam_review_rows = [
        {
            "日期": row.get("日期", ""),
            "好评率": row.get("热度值", ""),
            "好评数/总数": row.get("标题", ""),
            "游戏名": row.get("关键词", ""),
        }
        for row in data.trends
        if row.get("类型") == "Steam好评率(90天)"
    ]
    if steam_review_rows:
        sections.append(
            (
                "Steam好评率(90天)",
                sorted(steam_review_rows, key=lambda row: _series_sort_key(row.get("日期"))),
                ["日期", "好评率", "好评数/总数", "游戏名"],
                ["好评率"],
            )
        )

    if data.steam_player_peaks:
        steam_ccu_rows = [
            {
                "日期": row.get("日期", ""),
                "CCU Peak": row.get("在线峰值", ""),
                "时间戳(UTC)": row.get("时间戳(UTC)", ""),
                "游戏名": row.get("游戏名", ""),
                "App ID": row.get("App ID", ""),
            }
            for row in data.steam_player_peaks
        ]
        sections.append(
            (
                "Steam CCU Peak趋势",
                sorted(steam_ccu_rows, key=lambda row: _series_sort_key(row.get("日期"))),
                ["日期", "CCU Peak", "时间戳(UTC)", "游戏名", "App ID"],
                ["CCU Peak"],
            )
        )

    if data.google_trends:
        google_rows = [
            {
                "日期": row.get("日期", ""),
                "热度值": row.get("热度值", ""),
                "关键词": row.get("关键词", ""),
                "地区": row.get("地区", ""),
                "时间范围": row.get("时间范围", ""),
            }
            for row in data.google_trends
        ]
        sections.append(
            (
                "Google Trends搜索热度",
                sorted(google_rows, key=lambda row: _series_sort_key(row.get("日期"))),
                ["日期", "热度值", "关键词", "地区", "时间范围"],
                ["热度值"],
            )
        )

    if data.monitor_metrics:
        twitch_rows = [
            {
                "日期": row.get("日期", ""),
                "Twitch平均观看": row.get("Twitch平均观看", ""),
                "Twitch峰值观看": row.get("Twitch峰值观看", ""),
                "游戏名": row.get("游戏名", ""),
                "App ID": row.get("App ID", ""),
            }
            for row in data.monitor_metrics
            if row.get("Twitch平均观看") not in (None, "") or row.get("Twitch峰值观看") not in (None, "")
        ]
        if twitch_rows:
            sections.append(
                (
                    "Twitch Tracker观看趋势",
                    sorted(twitch_rows, key=lambda row: _series_sort_key(row.get("日期"))),
                    ["日期", "Twitch平均观看", "Twitch峰值观看", "游戏名", "App ID"],
                    ["Twitch平均观看", "Twitch峰值观看"],
                )
            )

    qimai_rows = [
        {
            "日期": row.get("日期", ""),
            "指标": row.get("指标", ""),
            "值": row.get("值", ""),
            "游戏名": row.get("游戏名", ""),
        }
        for row in data.trends
        if row.get("数据源") == "Qimai(AppStore)"
    ]
    if qimai_rows:
        sections.append(
            (
                "Qimai/AppStore趋势",
                sorted(qimai_rows, key=lambda row: (str(row.get("指标", "")), _series_sort_key(row.get("日期")))),
                ["日期", "指标", "值", "游戏名"],
                ["值"],
            )
        )

    other_trend_rows = [
        row
        for row in data.trends
        if row.get("类型") not in {"Steam好评率(90天)", "搜索热度"}
        and row.get("数据源") != "Qimai(AppStore)"
    ]
    if other_trend_rows:
        sections.append(
            (
                "其他趋势数据",
                sorted(other_trend_rows, key=lambda row: _series_sort_key(row.get("日期"))),
                _collect_headers(other_trend_rows),
                ["热度值", "值"],
            )
        )

    if not sections:
        if data.trends:
            _write_trends_sheet(wb, data.trends)
        return

    ws = wb.create_sheet(_unique_sheet_title(wb, "趋势数据"))
    current_row = 1
    for title, rows, headers, chart_value_headers in sections:
        current_row = _write_trend_section(
            ws,
            title=title,
            rows=rows,
            headers=headers,
            numeric_headers=chart_value_headers,
            start_row=current_row,
        )
        current_row += 3

    ws.freeze_panes = "A3"
    max_header_count = max((len(headers) for _, _, headers, _ in sections), default=1)
    _auto_column_width(ws, [f"col_{index}" for index in range(max_header_count)], max_width=28)


def _write_trend_section(
    ws,
    *,
    title: str,
    rows: list[dict[str, Any]],
    headers: list[str],
    numeric_headers: list[str],
    start_row: int,
) -> int:
    title_width = max(len(headers), 4)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=title_width)
    title_cell = ws.cell(row=start_row, column=1, value=f"{title}（{len(rows)}条）")
    title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_cell.font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(vertical="center", wrap_text=True)

    header_row = start_row + 1
    _write_header_row(ws, headers, row_num=header_row)
    for row_index, row in enumerate(rows, start=header_row + 1):
        for col_index, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=row.get(header, ""))
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT
            cell.border = DATA_BORDER
            if row_index % 2 == 0:
                cell.fill = ALT_ROW_FILL

    if len(rows) >= 2 and "日期" in headers:
        _add_section_line_chart(
            ws,
            title=title,
            headers=headers,
            numeric_headers=numeric_headers,
            header_row=header_row,
            row_count=len(rows),
        )

    return header_row + len(rows)


def _add_section_line_chart(
    ws,
    *,
    title: str,
    headers: list[str],
    numeric_headers: list[str],
    header_row: int,
    row_count: int,
) -> None:
    date_col = headers.index("日期") + 1
    value_cols = [
        headers.index(header) + 1
        for header in numeric_headers
        if header in headers and _column_has_numeric_values(ws, headers.index(header) + 1, header_row + 1, header_row + row_count)
    ]
    if not value_cols:
        return

    chart = LineChart()
    chart.title = title
    chart.y_axis.title = "Value"
    chart.x_axis.title = "日期"
    chart.style = 10
    chart.width = 22
    chart.height = 10
    for value_col in value_cols:
        data_ref = Reference(ws, min_col=value_col, min_row=header_row, max_row=header_row + row_count)
        chart.add_data(data_ref, titles_from_data=True)
    cats_ref = Reference(ws, min_col=date_col, min_row=header_row + 1, max_row=header_row + row_count)
    chart.set_categories(cats_ref)
    chart_anchor = f"{get_column_letter(len(headers) + 2)}{header_row}"
    ws.add_chart(chart, chart_anchor)


def _column_has_numeric_values(ws, column: int, start_row: int, end_row: int) -> bool:
    for row_index in range(start_row, end_row + 1):
        if isinstance(ws.cell(row=row_index, column=column).value, (int, float)):
            return True
    return False


def _write_trends_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    """写入趋势数据 Sheet 并生成折线图。"""
    ws = wb.create_sheet("趋势数据")
    headers = _collect_headers(rows)

    _write_header_row(ws, headers, row_num=1)

    for i, row in enumerate(rows, start=2):
        for j, header in enumerate(headers, start=1):
            cell = ws.cell(row=i, column=j, value=row.get(header, ""))
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT
            cell.border = DATA_BORDER
            if i % 2 == 0:
                cell.fill = ALT_ROW_FILL

    _auto_column_width(ws, headers)
    ws.freeze_panes = "A2"

    # 生成趋势折线图
    _add_trends_chart(ws, headers, rows)


def _write_related_queries_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    """写入相关搜索词 Sheet。"""
    ws = wb.create_sheet("相关搜索词")
    headers = _collect_headers(rows)

    _write_header_row(ws, headers, row_num=1)

    for i, row in enumerate(rows, start=2):
        for j, header in enumerate(headers, start=1):
            cell = ws.cell(row=i, column=j, value=row.get(header, ""))
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT
            cell.border = DATA_BORDER
            if i % 2 == 0:
                cell.fill = ALT_ROW_FILL

    _auto_column_width(ws, headers)
    ws.freeze_panes = "A2"

    # 添加相关搜索词条形图
    _add_related_queries_chart(ws, headers, rows)


def _write_llm_sheet(wb: Workbook, content: str, title: str) -> None:
    """写入 LLM 分析报告 Sheet。"""
    ws = wb.create_sheet("AI 分析报告")

    # 标题
    ws["A1"] = title
    ws["A1"].font = Font(name="微软雅黑", bold=True, size=14, color="2F5496")
    ws.merge_cells("A1:H1")

    # 内容 —— 按段落分行写入
    paragraphs = content.split("\n")
    current_row = 3
    for para in paragraphs:
        para = para.strip()
        if not para:
            current_row += 1
            continue

        cell = ws.cell(row=current_row, column=1, value=para)

        # Markdown 标题检测
        if para.startswith("# "):
            cell.value = para.lstrip("# ")
            cell.font = Font(name="微软雅黑", bold=True, size=14, color="2F5496")
        elif para.startswith("## "):
            cell.value = para.lstrip("# ")
            cell.font = Font(name="微软雅黑", bold=True, size=12, color="2F5496")
        elif para.startswith("### "):
            cell.value = para.lstrip("# ")
            cell.font = Font(name="微软雅黑", bold=True, size=11)
        elif para.startswith("- ") or para.startswith("* "):
            cell.font = Font(name="微软雅黑", size=10)
            cell.alignment = Alignment(indent=2)
        else:
            cell.font = Font(name="微软雅黑", size=10)

        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
        current_row += 1

    ws.column_dimensions["A"].width = 120


# ==================== 图表生成 ====================

def _add_overview_chart(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    """在概览 Sheet 中生成柱状图（当前在线/评论总量对比）。"""
    if len(rows) < 1:
        return

    # 查找数值列
    numeric_cols = []
    for h in ["当前在线", "评论总量"]:
        if h in headers:
            numeric_cols.append(h)

    if not numeric_cols:
        return

    # 构建辅助数据区域（在数据区域右侧）
    chart_start_col = len(headers) + 3
    chart_start_row = 1

    # 写图表标题行
    ws.cell(row=chart_start_row, column=chart_start_col, value="游戏名")
    for k, col_name in enumerate(numeric_cols, start=1):
        ws.cell(row=chart_start_row, column=chart_start_col + k, value=col_name)

    for i, row in enumerate(rows, start=1):
        ws.cell(row=chart_start_row + i, column=chart_start_col, value=row.get("游戏名", ""))
        for k, col_name in enumerate(numeric_cols, start=1):
            val = row.get(col_name, 0)
            ws.cell(row=chart_start_row + i, column=chart_start_col + k, value=val if isinstance(val, (int, float)) else 0)

    # 生成柱状图
    chart = BarChart()
    chart.type = "col"
    chart.title = "游戏核心指标对比"
    chart.y_axis.title = "数值"
    chart.style = 10
    chart.width = 20
    chart.height = 12

    data_ref = Reference(
        ws,
        min_col=chart_start_col + 1,
        max_col=chart_start_col + len(numeric_cols),
        min_row=chart_start_row,
        max_row=chart_start_row + len(rows),
    )
    cats_ref = Reference(
        ws,
        min_col=chart_start_col,
        min_row=chart_start_row + 1,
        max_row=chart_start_row + len(rows),
    )
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    ws.add_chart(chart, f"A{len(rows) + 4}")


def _add_trends_chart(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    """在趋势 Sheet 中生成折线图。"""
    # 过滤出搜索热度类型的数据
    heat_rows = [r for r in rows if r.get("类型") == "搜索热度" and isinstance(r.get("热度值"), (int, float))]

    if len(heat_rows) < 2:
        return

    # 查找日期和热度值列索引
    if "日期" not in headers or "热度值" not in headers:
        return

    date_col = headers.index("日期") + 1
    value_col = headers.index("热度值") + 1

    chart = LineChart()
    chart.title = "搜索热度趋势"
    chart.y_axis.title = "热度 (0-100)"
    chart.x_axis.title = "日期"
    chart.style = 10
    chart.width = 28
    chart.height = 14

    # 找到第一个和最后一个搜索热度行
    first_heat_row = None
    last_heat_row = None
    for i, row in enumerate(rows, start=2):
        if row.get("类型") == "搜索热度":
            if first_heat_row is None:
                first_heat_row = i
            last_heat_row = i

    if first_heat_row is None or last_heat_row is None:
        return

    data_ref = Reference(ws, min_col=value_col, min_row=first_heat_row - 1, max_row=last_heat_row)
    cats_ref = Reference(ws, min_col=date_col, min_row=first_heat_row, max_row=last_heat_row)

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    ws.add_chart(chart, f"A{len(rows) + 4}")


def _add_related_queries_chart(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    """在相关搜索词 Sheet 中生成水平条形图。"""
    # 只取 top 10 热门词
    top_rows = [r for r in rows if r.get("类型") == "热门"][:10]
    if len(top_rows) < 2:
        return

    # 在数据右侧构建辅助区域
    chart_start_col = len(headers) + 3
    chart_start_row = 1

    ws.cell(row=chart_start_row, column=chart_start_col, value="查询词")
    ws.cell(row=chart_start_row, column=chart_start_col + 1, value="热度值")

    for i, row in enumerate(top_rows, start=1):
        ws.cell(row=chart_start_row + i, column=chart_start_col, value=row.get("查询词", ""))
        val = row.get("热度值", 0)
        ws.cell(row=chart_start_row + i, column=chart_start_col + 1, value=val if isinstance(val, (int, float)) else 0)

    chart = BarChart()
    chart.type = "bar"
    chart.title = "热门相关搜索词 Top 10"
    chart.y_axis.title = "查询词"
    chart.x_axis.title = "热度值"
    chart.style = 10
    chart.width = 22
    chart.height = 14

    data_ref = Reference(ws, min_col=chart_start_col + 1, min_row=chart_start_row, max_row=chart_start_row + len(top_rows))
    cats_ref = Reference(ws, min_col=chart_start_col, min_row=chart_start_row + 1, max_row=chart_start_row + len(top_rows))

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    ws.add_chart(chart, f"A{len(rows) + 4}")


# ==================== 工具函数 ====================

def _collect_headers(rows: list[dict[str, Any]]) -> list[str]:
    """从多行数据中收集所有出现过的列名（保持插入顺序）。"""
    seen: set[str] = set()
    headers: list[str] = []
    for row in rows:
        for key in row:
            if key not in seen:
                seen.add(key)
                headers.append(key)
    return headers


def _series_sort_key(value: Any) -> tuple[int, int, int, str]:
    text = str(value or "").strip()
    parsed = _parse_series_date(text)
    if parsed:
        return (0, parsed.year, parsed.month, text)
    return (1, 9999, 12, text)


def _parse_series_date(text: str) -> datetime | None:
    if not text:
        return None
    lowered = text.lower().strip()
    if lowered in {"last 30 days", "last 7 days", "recent"}:
        return datetime(datetime.now().year, datetime.now().month, 1)

    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m", "%Y/%m", "%b %Y", "%B %Y"):
        try:
            return datetime.strptime(text[:10] if fmt in {"%Y-%m-%d", "%Y/%m/%d"} else text, fmt)
        except ValueError:
            continue

    match = re.match(r"^\s*(\d{4})年\s*(\d{1,2})月", text)
    if match:
        return datetime(int(match.group(1)), int(match.group(2)), 1)
    match = re.match(r"^\s*(\d{1,2})月\s*(\d{4})年?", text)
    if match:
        return datetime(int(match.group(2)), int(match.group(1)), 1)
    return None


def _write_header_row(ws, headers: list[str], row_num: int = 1) -> None:
    """写入带样式的标题行。"""
    for j, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_num, column=j, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = HEADER_BORDER


def _auto_column_width(ws, headers: list[str], max_width: int = 40) -> None:
    """根据内容自动调整列宽。"""
    for j, header in enumerate(headers, start=1):
        col_letter = get_column_letter(j)
        # 取标题宽度和内容最大宽度中的较大值
        header_width = len(header) * 2.2  # CJK 字符加权
        content_width = 0
        for row in ws.iter_rows(min_row=2, min_col=j, max_col=j):
            for cell in row:
                if cell.value:
                    val_str = str(cell.value)
                    # CJK 字符宽度约等于 2 个英文字符
                    w = sum(2 if ord(c) > 127 else 1 for c in val_str[:50])
                    content_width = max(content_width, w)

        width = min(max(header_width, content_width + 2), max_width)
        ws.column_dimensions[col_letter].width = max(width, 8)
