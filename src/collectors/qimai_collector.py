"""
Qimai collector.

The collector uses Playwright with a persistent browser profile because Qimai
serves most useful data through dynamic `api.qimai.cn` requests that depend on
its logged-in web session and request signing.
"""

from __future__ import annotations

import asyncio
import csv
import os
import random
import re
import statistics
import sys
import time
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Iterable

from loguru import logger

from src.collectors.base import BaseCollector, CollectResult, CollectTarget
from src.core.config import get as get_config
from src.core.registry import registry

try:
    from playwright.async_api import Response as AsyncResponse
    from playwright.async_api import async_playwright
    from playwright.sync_api import Response as SyncResponse
    from playwright.sync_api import sync_playwright
except ImportError:  # pragma: no cover - dependency is optional at import time
    async_playwright = None
    sync_playwright = None
    AsyncResponse = Any
    SyncResponse = Any


class QimaiScrapeFailed(Exception):
    """Raised when Qimai scraping fails."""


@registry.register("collector", "qimai")
class QimaiCollector(BaseCollector):
    """
    Collect Qimai iOS App Store metrics.

    Target params:
      - qimai_app_id or app_id: Qimai/App Store app id.
      - country: defaults to cn.

    Requested output fields:
      - grossing_rank_cn
      - ios_grossing_rank_trend
      - appstore_rating_cn
      - appstore_review_trend
      - dau_avg_30d
      - dau_trend_90d
      - downloads_avg_30d
      - downloads_trend_90d
      - revenue_avg_30d
      - revenue_trend_90d
    """

    BASE_URL = "https://www.qimai.cn"

    def __init__(self, config: dict[str, Any] | None = None):
        super().__init__(config)
        self._headless = self.config.get("headless", get_config("qimai.headless", True))
        self._timeout = int(self.config.get("timeout", get_config("qimai.timeout", 45000)))
        self._delay = float(self.config.get("request_delay", get_config("qimai.request_delay", 8.0)))
        self._jitter = float(self.config.get("request_jitter", get_config("qimai.request_jitter", 4.0)))
        self._click_delay = float(self.config.get("click_delay", get_config("qimai.click_delay", 1.5)))
        self._scroll_delay = float(self.config.get("scroll_delay", get_config("qimai.scroll_delay", 2.0)))
        self._user_data_dir = self.config.get(
            "user_data_dir",
            get_config("qimai.user_data_dir", os.path.join(os.getcwd(), "data", "qimai_profile")),
        )
        self._max_api_payloads = int(self.config.get("max_api_payloads", get_config("qimai.max_api_payloads", 80)))
        self._cdp_enabled = bool(self.config.get("cdp_enabled", get_config("qimai.cdp_enabled", True)))
        self._cdp_port = int(self.config.get("cdp_port", get_config("qimai.cdp_port", 9222)))
        self._cdp_required = bool(self.config.get("cdp_required", get_config("qimai.cdp_required", False)))

    async def setup(self, config: dict[str, Any] | None = None) -> None:
        await super().setup(config)
        logger.info(
            "[QimaiCollector] initialized user-data-dir={} delay={} jitter={} click_delay={} scroll_delay={}",
            self._user_data_dir,
            self._delay,
            self._jitter,
            self._click_delay,
            self._scroll_delay,
        )

    async def collect(self, target: CollectTarget) -> CollectResult:
        app_id = _normalize_app_id(target.params.get("qimai_app_id", target.params.get("app_id", "")))
        if not app_id:
            return CollectResult(
                target=target,
                success=False,
                error="Qimai requires 'qimai_app_id' or 'app_id'",
            )

        country = str(target.params.get("country") or "cn").lower()
        logger.info("[Qimai] collecting {} app_id={} country={}", target.name, app_id, country)

        try:
            if self._should_use_threaded_playwright():
                qimai_data = await asyncio.to_thread(self._scrape_sync, app_id, country)
            else:
                qimai_data = await self._scrape_async(app_id, country)
        except Exception as exc:
            logger.exception("[Qimai] collection failed")
            return CollectResult(
                target=target,
                success=False,
                error=f"Qimai collection failed: {exc}",
                metadata={"collector": "qimai"},
            )

        app_name = qimai_data.get("app_name") or target.name
        if _looks_like_navigation_text(app_name):
            app_name = target.name
        merged_data = {
            "collector": "qimai",
            "game_name": app_name,
            "app_id": app_id,
            "qimai": qimai_data,
            "snapshot": {
                "name": app_name,
                "review_score": qimai_data.get("rating", qimai_data.get("appstore_rating_cn", "")),
                "total_reviews": qimai_data.get("rating_count", 0),
                "free_rank": qimai_data.get("free_rank", ""),
                "grossing_rank": qimai_data.get("grossing_rank", qimai_data.get("grossing_rank_cn", "")),
                "grossing_rank_cn": qimai_data.get("grossing_rank_cn", ""),
                "appstore_rating_cn": qimai_data.get("appstore_rating_cn", ""),
                "dau_avg_30d": qimai_data.get("dau_avg_30d"),
                "downloads_avg_30d": qimai_data.get("downloads_avg_30d"),
                "revenue_avg_30d": qimai_data.get("revenue_avg_30d"),
            },
        }

        return CollectResult(
            target=target,
            data=merged_data,
            success=True,
            metadata={"collector": "qimai"},
        )

    def _should_use_threaded_playwright(self) -> bool:
        if sys.platform != "win32":
            return False
        loop_name = asyncio.get_running_loop().__class__.__name__
        return "Selector" in loop_name

    def _urls(self, app_id: str, country: str) -> list[tuple[str, str]]:
        return [
            ("baseinfo", f"{self.BASE_URL}/app/baseinfo/appid/{app_id}/country/{country}"),
            ("rank", f"{self.BASE_URL}/app/rank/appid/{app_id}/country/{country}"),
            ("comment", f"{self.BASE_URL}/app/comment/appid/{app_id}/country/{country}"),
            ("appstatus", f"{self.BASE_URL}/app/appstatus/appid/{app_id}/country/{country}"),
            ("download", f"{self.BASE_URL}/app/download/appid/{app_id}/country/{country}"),
            ("income", f"{self.BASE_URL}/app/income/appid/{app_id}/country/{country}"),
        ]

    async def _scrape_async(self, app_id: str, country: str) -> dict[str, Any]:
        if async_playwright is None:
            raise QimaiScrapeFailed("playwright is not installed")

        state = _QimaiCaptureState(app_id=app_id, country=country, max_payloads=self._max_api_payloads)

        async with async_playwright() as p:
            browser = None
            context = None
            is_cdp = False
            if self._cdp_enabled:
                try:
                    browser = await p.chromium.connect_over_cdp(f"http://127.0.0.1:{self._cdp_port}")
                    context = browser.contexts[0] if browser.contexts else await browser.new_context(accept_downloads=True)
                    is_cdp = True
                    logger.info("[Qimai] Connected to local browser over CDP port {}", self._cdp_port)
                except Exception as exc:
                    logger.warning("[Qimai] CDP connection failed, falling back to persistent profile: {}", exc)
                    if self._cdp_required:
                        raise QimaiScrapeFailed(f"CDP connection failed on port {self._cdp_port}: {exc}") from exc
            if context is None:
                context = await p.chromium.launch_persistent_context(
                    self._user_data_dir,
                    headless=self._headless,
                    accept_downloads=True,
                    args=["--disable-blink-features=AutomationControlled"],
                    ignore_default_args=["--enable-automation"],
                    viewport={"width": 1920, "height": 1080},
                    user_agent=_desktop_user_agent(),
                )
            try:
                page = None
                on_response = None
                page = await context.new_page()
                await page.add_init_script(_stealth_script())
                capture_tasks: list[asyncio.Task[Any]] = []

                def on_response(response: Any) -> None:
                    if not state.accepting_responses:
                        return
                    capture_tasks.append(asyncio.create_task(state.capture_async(response)))

                page.on("response", on_response)

                for index, (page_name, url) in enumerate(self._urls(app_id, country)):
                    if index:
                        await self._polite_wait_async(f"before navigating {page_name}")
                    await page.goto(url, wait_until="domcontentloaded", timeout=self._timeout)
                    await self._configure_page_async(page, page_name)
                    await self._polite_wait_async(f"after configuring {page_name}")
                    await self._capture_visible_chart_async(page, state, page_name)
                    await self._capture_page_export_async(page, state, page_name)
                    await _safe_scroll_async(page)
                    await self._scroll_wait_async(page_name)
                    state.add_page_text(page_name, await page.locator("body").inner_text(timeout=5000))
                    state.add_page_url(page_name, page.url)
                await self._capture_direct_metrics_async(page, state, app_id, country)
                await self._capture_pred_estimates_async(page, state, app_id, country)
                if country == "cn":
                    await self._capture_public_rank_async(page, state)
                if capture_tasks:
                    await asyncio.gather(*capture_tasks, return_exceptions=True)
            finally:
                state.stop_capture()
                try:
                    if page is not None and on_response is not None:
                        page.remove_listener("response", on_response)
                except Exception:
                    pass
                if "capture_tasks" in locals() and capture_tasks:
                    await asyncio.gather(*capture_tasks, return_exceptions=True)
                if not is_cdp:
                    await context.close()

        return state.build_result()

    def _scrape_sync(self, app_id: str, country: str) -> dict[str, Any]:
        if sync_playwright is None:
            raise QimaiScrapeFailed("playwright is not installed")

        state = _QimaiCaptureState(app_id=app_id, country=country, max_payloads=self._max_api_payloads)

        with sync_playwright() as p:
            browser = None
            context = None
            is_cdp = False
            if self._cdp_enabled:
                try:
                    browser = p.chromium.connect_over_cdp(f"http://127.0.0.1:{self._cdp_port}")
                    context = browser.contexts[0] if browser.contexts else browser.new_context(accept_downloads=True)
                    is_cdp = True
                    logger.info("[Qimai] Connected to local browser over CDP port {}", self._cdp_port)
                except Exception as exc:
                    logger.warning("[Qimai] CDP connection failed, falling back to persistent profile: {}", exc)
                    if self._cdp_required:
                        raise QimaiScrapeFailed(f"CDP connection failed on port {self._cdp_port}: {exc}") from exc
            if context is None:
                context = p.chromium.launch_persistent_context(
                    self._user_data_dir,
                    headless=self._headless,
                    accept_downloads=True,
                    args=["--disable-blink-features=AutomationControlled"],
                    ignore_default_args=["--enable-automation"],
                    viewport={"width": 1920, "height": 1080},
                    user_agent=_desktop_user_agent(),
                )
            try:
                page = None
                page = context.new_page()
                page.add_init_script(_stealth_script())
                page.on("response", state.capture_sync)

                for index, (page_name, url) in enumerate(self._urls(app_id, country)):
                    if index:
                        self._polite_wait_sync(f"before navigating {page_name}")
                    page.goto(url, wait_until="domcontentloaded", timeout=self._timeout)
                    self._configure_page_sync(page, page_name)
                    self._polite_wait_sync(f"after configuring {page_name}")
                    self._capture_visible_chart_sync(page, state, page_name)
                    self._capture_page_export_sync(page, state, page_name)
                    _safe_scroll_sync(page)
                    self._scroll_wait_sync(page_name)
                    state.add_page_text(page_name, page.locator("body").inner_text(timeout=5000))
                    state.add_page_url(page_name, page.url)
                self._capture_direct_metrics_sync(page, state, app_id, country)
                self._capture_pred_estimates_sync(page, state, app_id, country)
                if country == "cn":
                    self._capture_public_rank_sync(page, state)
            finally:
                state.stop_capture()
                try:
                    if page is not None:
                        page.remove_listener("response", state.capture_sync)
                except Exception:
                    pass
                if not is_cdp:
                    context.close()

        return state.build_result()

    async def _configure_page_async(self, page: Any, page_name: str) -> None:
        if page_name == "rank":
            await _click_filter_option_async(page, "\u8bbe\u5907", "iPhone", self._click_delay, self._jitter)
            await _click_filter_option_async(page, "\u699c\u5355\u7c7b\u578b", "\u7545\u9500", self._click_delay, self._jitter)
            await _click_filter_option_async(page, "\u65e5\u671f", "\u8fd1\u4e09\u4e2a\u6708", self._click_delay, self._jitter)
        elif page_name == "comment":
            await _click_filter_option_async(page, "\u7edf\u8ba1\u65b9\u5f0f", "\u6bcf\u65e5\u53d8\u52a8", self._click_delay, self._jitter)
            await _click_filter_option_async(page, "\u65f6\u95f4", "\u8fd1\u4e09\u4e2a\u6708", self._click_delay, self._jitter)
        elif page_name in {"appstatus", "download", "income"}:
            await _click_filter_option_async(page, "\u8bbe\u5907", "iPhone", self._click_delay, self._jitter)
            await _click_filter_option_async(page, "\u65e5\u671f", "\u8fd1\u4e09\u4e2a\u6708", self._click_delay, self._jitter)
        if page_name == "appstatus":
            for label in ("DAU", "\u65e5\u6d3b", "\u6d3b\u8dc3"):
                await _click_text_async(page, label, self._click_delay, self._jitter)

    def _configure_page_sync(self, page: Any, page_name: str) -> None:
        if page_name == "rank":
            _click_filter_option_sync(page, "\u8bbe\u5907", "iPhone", self._click_delay, self._jitter)
            _click_filter_option_sync(page, "\u699c\u5355\u7c7b\u578b", "\u7545\u9500", self._click_delay, self._jitter)
            _click_filter_option_sync(page, "\u65e5\u671f", "\u8fd1\u4e09\u4e2a\u6708", self._click_delay, self._jitter)
        elif page_name == "comment":
            _click_filter_option_sync(page, "\u7edf\u8ba1\u65b9\u5f0f", "\u6bcf\u65e5\u53d8\u52a8", self._click_delay, self._jitter)
            _click_filter_option_sync(page, "\u65f6\u95f4", "\u8fd1\u4e09\u4e2a\u6708", self._click_delay, self._jitter)
        elif page_name in {"appstatus", "download", "income"}:
            _click_filter_option_sync(page, "\u8bbe\u5907", "iPhone", self._click_delay, self._jitter)
            _click_filter_option_sync(page, "\u65e5\u671f", "\u8fd1\u4e09\u4e2a\u6708", self._click_delay, self._jitter)
        if page_name == "appstatus":
            for label in ("DAU", "\u65e5\u6d3b", "\u6d3b\u8dc3"):
                _click_text_sync(page, label, self._click_delay, self._jitter)

    async def _polite_wait_async(self, reason: str) -> None:
        delay = _throttle_seconds(self._delay, self._jitter)
        logger.debug("[Qimai] throttle {}: {:.2f}s", reason, delay)
        await asyncio.sleep(delay)

    def _polite_wait_sync(self, reason: str) -> None:
        delay = _throttle_seconds(self._delay, self._jitter)
        logger.debug("[Qimai] throttle {}: {:.2f}s", reason, delay)
        time.sleep(delay)

    async def _scroll_wait_async(self, page_name: str) -> None:
        delay = _throttle_seconds(self._scroll_delay, min(self._jitter, self._scroll_delay))
        logger.debug("[Qimai] throttle after scroll {}: {:.2f}s", page_name, delay)
        await asyncio.sleep(delay)

    def _scroll_wait_sync(self, page_name: str) -> None:
        delay = _throttle_seconds(self._scroll_delay, min(self._jitter, self._scroll_delay))
        logger.debug("[Qimai] throttle after scroll {}: {:.2f}s", page_name, delay)
        time.sleep(delay)

    async def _capture_public_rank_async(self, page: Any, state: "_QimaiCaptureState") -> None:
        url = f"{self.BASE_URL}/rank/index/brand/grossing/device/iphone/country/{state.country}/genre/36"
        try:
            await self._polite_wait_async("before navigating public rank")
            await page.goto(url, wait_until="domcontentloaded", timeout=self._timeout)
            await _click_filter_option_async(page, "\u8bbe\u5907", "iPhone", self._click_delay, self._jitter)
            await _click_filter_option_async(page, "\u699c\u5355\u7c7b\u578b", "\u7545\u9500", self._click_delay, self._jitter)
            await _safe_scroll_async(page)
            await self._scroll_wait_async("public_rank")
            state.add_page_text("public_rank", await page.locator("body").inner_text(timeout=8000))
            state.add_page_url("public_rank", page.url)
        except Exception as exc:
            logger.debug("[Qimai] public rank fallback skipped: {}", exc)

    def _capture_public_rank_sync(self, page: Any, state: "_QimaiCaptureState") -> None:
        url = f"{self.BASE_URL}/rank/index/brand/grossing/device/iphone/country/{state.country}/genre/36"
        try:
            self._polite_wait_sync("before navigating public rank")
            page.goto(url, wait_until="domcontentloaded", timeout=self._timeout)
            _click_filter_option_sync(page, "\u8bbe\u5907", "iPhone", self._click_delay, self._jitter)
            _click_filter_option_sync(page, "\u699c\u5355\u7c7b\u578b", "\u7545\u9500", self._click_delay, self._jitter)
            _safe_scroll_sync(page)
            self._scroll_wait_sync("public_rank")
            state.add_page_text("public_rank", page.locator("body").inner_text(timeout=8000))
            state.add_page_url("public_rank", page.url)
        except Exception as exc:
            logger.debug("[Qimai] public rank fallback skipped: {}", exc)

    async def _capture_pred_estimates_async(self, page: Any, state: "_QimaiCaptureState", app_id: str, country: str) -> None:
        try:
            await self._polite_wait_async("before qimai pred estimates")
            responses = await page.evaluate(_qimai_pred_estimate_script(), _qimai_pred_params(app_id, country))
            _append_pred_estimate_payloads(state, responses)
        except Exception as exc:
            state.warnings.append(f"Qimai pred estimate request failed: {exc}")
            logger.debug("[Qimai] pred estimate request failed: {}", exc)

    def _capture_pred_estimates_sync(self, page: Any, state: "_QimaiCaptureState", app_id: str, country: str) -> None:
        try:
            self._polite_wait_sync("before qimai pred estimates")
            responses = page.evaluate(_qimai_pred_estimate_script(), _qimai_pred_params(app_id, country))
            _append_pred_estimate_payloads(state, responses)
        except Exception as exc:
            state.warnings.append(f"Qimai pred estimate request failed: {exc}")
            logger.debug("[Qimai] pred estimate request failed: {}", exc)

    async def _capture_direct_metrics_async(self, page: Any, state: "_QimaiCaptureState", app_id: str, country: str) -> None:
        try:
            await self._polite_wait_async("before qimai direct metrics")
            responses = await page.evaluate(_qimai_direct_metrics_script(), _qimai_direct_metric_params(app_id, country))
            _append_direct_metric_payloads(state, responses)
        except Exception as exc:
            state.warnings.append(f"Qimai direct metric request failed: {exc}")
            logger.debug("[Qimai] direct metric request failed: {}", exc)

    def _capture_direct_metrics_sync(self, page: Any, state: "_QimaiCaptureState", app_id: str, country: str) -> None:
        try:
            self._polite_wait_sync("before qimai direct metrics")
            responses = page.evaluate(_qimai_direct_metrics_script(), _qimai_direct_metric_params(app_id, country))
            _append_direct_metric_payloads(state, responses)
        except Exception as exc:
            state.warnings.append(f"Qimai direct metric request failed: {exc}")
            logger.debug("[Qimai] direct metric request failed: {}", exc)

    async def _capture_visible_chart_async(self, page: Any, state: "_QimaiCaptureState", page_name: str) -> None:
        if page_name not in {"rank", "comment", "appstatus", "download", "income"}:
            return
        try:
            series = await page.evaluate(_qimai_echarts_extract_script(), page_name)
            state.add_chart_series(page_name, series)
        except Exception as exc:
            logger.debug("[Qimai] visible chart capture skipped for {}: {}", page_name, exc)

    def _capture_visible_chart_sync(self, page: Any, state: "_QimaiCaptureState", page_name: str) -> None:
        if page_name not in {"rank", "comment", "appstatus", "download", "income"}:
            return
        try:
            series = page.evaluate(_qimai_echarts_extract_script(), page_name)
            state.add_chart_series(page_name, series)
        except Exception as exc:
            logger.debug("[Qimai] visible chart capture skipped for {}: {}", page_name, exc)

    async def _capture_page_export_async(self, page: Any, state: "_QimaiCaptureState", page_name: str) -> None:
        if page_name not in {"rank", "comment", "download", "income"}:
            return
        try:
            await _click_table_view_async(page)
            async with page.expect_download(timeout=3500) as download_info:
                clicked = await _click_export_button_async(page)
                if not clicked:
                    return
            download = await download_info.value
            export_path = await download.path()
            rows = _read_qimai_export_rows(export_path, download.suggested_filename)
            state.add_export_rows(page_name, rows, download.suggested_filename)
        except Exception as exc:
            logger.debug("[Qimai] export capture skipped for {}: {}", page_name, exc)

    def _capture_page_export_sync(self, page: Any, state: "_QimaiCaptureState", page_name: str) -> None:
        if page_name not in {"rank", "comment", "download", "income"}:
            return
        try:
            _click_table_view_sync(page)
            with page.expect_download(timeout=3500) as download_info:
                clicked = _click_export_button_sync(page)
                if not clicked:
                    return
            download = download_info.value
            rows = _read_qimai_export_rows(download.path(), download.suggested_filename)
            state.add_export_rows(page_name, rows, download.suggested_filename)
        except Exception as exc:
            logger.debug("[Qimai] export capture skipped for {}: {}", page_name, exc)


class _QimaiCaptureState:
    def __init__(self, *, app_id: str, country: str, max_payloads: int):
        self.app_id = app_id
        self.country = country
        self.max_payloads = max_payloads
        self.accepting_responses = True
        self.page_texts: dict[str, str] = {}
        self.page_urls: dict[str, str] = {}
        self.api_payloads: list[dict[str, Any]] = []
        self.chart_series: dict[str, list[dict[str, Any]]] = {}
        self.export_rows: dict[str, list[dict[str, Any]]] = {}
        self.warnings: list[str] = []

    def stop_capture(self) -> None:
        self.accepting_responses = False

    def add_page_text(self, page_name: str, text: str) -> None:
        self.page_texts[page_name] = text or ""
        if "\u5f53\u524d\u7f51\u7edc\u6216\u8d26\u53f7\u5f02\u5e38" in self.page_texts[page_name]:
            self.warnings.append("Qimai page reported network/account anomaly.")
        if "\u60a8\u8bbf\u95ee\u7684\u9875\u9762\u4e0d\u5b58\u5728" in self.page_texts[page_name]:
            self.warnings.append(f"Qimai page not found: {page_name}")

    def add_page_url(self, page_name: str, url: str) -> None:
        self.page_urls[page_name] = url

    def add_export_rows(self, page_name: str, rows: list[dict[str, Any]], filename: str = "") -> None:
        if not rows:
            return
        self.export_rows[page_name] = rows
        logger.info("[Qimai] captured {} export rows for {} from {}", len(rows), page_name, filename or "download")

    def add_chart_series(self, page_name: str, series: Any) -> None:
        if not isinstance(series, list):
            return
        normalized = _normalize_series(series)
        if not normalized:
            return
        self.chart_series[page_name] = normalized
        logger.info("[Qimai] captured {} visible chart points for {}", len(normalized), page_name)

    async def capture_async(self, response: AsyncResponse) -> None:
        try:
            await self._capture_response(response, is_async=True)
        except asyncio.CancelledError as exc:
            logger.debug("[Qimai] async response capture cancelled: {}", exc)
        except Exception as exc:
            if _is_ignorable_playwright_capture_error(exc):
                logger.debug("[Qimai] async response capture skipped after page close: {}", exc)
                return
            logger.debug("[Qimai] async response capture skipped: {}", exc)
        except BaseException as exc:
            if _is_ignorable_playwright_capture_error(exc):
                logger.debug("[Qimai] async response capture skipped after page close: {}", exc)
                return
            raise

    def capture_sync(self, response: SyncResponse) -> None:
        try:
            self._capture_response_sync(response)
        except asyncio.CancelledError as exc:
            logger.debug("[Qimai] response capture cancelled: {}", exc)
        except Exception as exc:
            if _is_ignorable_playwright_capture_error(exc):
                logger.debug("[Qimai] response capture skipped after page close: {}", exc)
                return
            logger.debug("[Qimai] response capture skipped: {}", exc)
        except BaseException as exc:
            if _is_ignorable_playwright_capture_error(exc):
                logger.debug("[Qimai] response capture skipped after page close: {}", exc)
                return
            raise

    async def _capture_response(self, response: Any, *, is_async: bool) -> None:
        try:
            if not self.accepting_responses:
                return
            if "api.qimai.cn" not in response.url or len(self.api_payloads) >= self.max_payloads:
                return
            content_type = response.headers.get("content-type", "")
            if "json" not in content_type and not response.url.endswith(".json"):
                return
            payload = await response.json() if is_async else response.json()
            self._append_payload(response.url, payload)
        except asyncio.CancelledError:
            raise
        except Exception as exc:
            logger.debug("[Qimai] async response capture skipped: {}", exc)

    def _capture_response_sync(self, response: Any) -> None:
        if not self.accepting_responses:
            return
        if "api.qimai.cn" not in response.url or len(self.api_payloads) >= self.max_payloads:
            return
        content_type = response.headers.get("content-type", "")
        if "json" not in content_type and not response.url.endswith(".json"):
            return
        self._append_payload(response.url, response.json())

    def _append_payload(self, url: str, payload: Any) -> None:
        if not isinstance(payload, (dict, list)):
            return
        if isinstance(payload, dict):
            code = payload.get("code")
            if code not in (None, 0, "0", 10000, "10000"):
                message = payload.get("msg") or payload.get("message") or "unknown error"
                self.warnings.append(f"Qimai API skipped non-success response code={code}: {message}")
                return
        self.api_payloads.append({"url": url, "payload": payload})

    def build_result(self) -> dict[str, Any]:
        result: dict[str, Any] = {
            "app_id": self.app_id,
            "country": self.country,
            "pages_crawled": self.page_urls,
            "api_response_count": len(self.api_payloads),
        }

        self._extract_visible_base_fields(result)
        self._extract_api_fields(result)
        self._extract_chart_fields(result)
        self._extract_export_fields(result)
        self._finalize_derived_fields(result)

        missing = [
            key
            for key in (
                "grossing_rank_cn",
                "ios_grossing_rank_trend",
                "appstore_rating_cn",
                "appstore_review_trend",
                "dau_avg_30d",
                "dau_trend_90d",
                "downloads_avg_30d",
                "downloads_trend_90d",
                "revenue_avg_30d",
                "revenue_trend_90d",
            )
            if result.get(key) in (None, "", [])
        ]
        if missing:
            self.warnings.append(
                "Missing fields may require a logged-in Qimai session or paid permission: "
                + ", ".join(missing)
            )
        if self.warnings:
            result["extraction_warnings"] = list(dict.fromkeys(self.warnings))
        return result

    def _extract_visible_base_fields(self, result: dict[str, Any]) -> None:
        base_text = self.page_texts.get("baseinfo", "")
        all_text = "\n".join(self.page_texts.values())

        app_name = _first_nonempty_line(base_text)
        if app_name and not _looks_like_navigation_text(app_name):
            result["app_name"] = app_name

        rating, rating_count = _extract_rating(base_text or all_text)
        if rating is not None:
            result["rating"] = rating
            result["appstore_rating_cn"] = rating
        if rating_count is not None:
            result["rating_count"] = rating_count

        free_rank = _extract_rank_near_label(all_text, "\u514d\u8d39")
        grossing_rank = _extract_rank_near_label(all_text, "\u7545\u9500")
        rank_page_grossing = _extract_grossing_rank_from_public_rank(
            self.page_texts.get("public_rank", ""),
            result.get("app_name", ""),
        )
        if free_rank:
            result["free_rank"] = free_rank
        if grossing_rank or rank_page_grossing:
            result["grossing_rank"] = rank_page_grossing or grossing_rank
            result["grossing_rank_cn"] = rank_page_grossing or grossing_rank
            if rank_page_grossing:
                result["grossing_rank_source"] = "qimai_public_rank_page"

        yesterday_downloads = _extract_yesterday_downloads(all_text)
        if yesterday_downloads is not None:
            result["yesterday_downloads"] = yesterday_downloads

    def _extract_api_fields(self, result: dict[str, Any]) -> None:
        all_payloads = [item["payload"] for item in self.api_payloads]
        joined_by_url = "\n".join(item["url"].lower() for item in self.api_payloads)

        scalar_candidates = list(_walk_json(all_payloads))
        if "appstore_rating_cn" not in result:
            rating = _find_scalar(scalar_candidates, ("rating", "score"), prefer_float=True)
            if rating is not None:
                result["rating"] = rating
                result["appstore_rating_cn"] = rating
        if "rating_count" not in result:
            count = _find_scalar(scalar_candidates, ("rating_count", "comment_num", "commentnum", "comment"))
            if count is not None:
                result["rating_count"] = count

        series_candidates = _extract_series_candidates(self.api_payloads)
        result.setdefault("raw_series_candidates", _series_debug_summary(series_candidates))

        if not result.get("ios_grossing_rank_trend"):
            result["ios_grossing_rank_trend"] = _pick_series(
                series_candidates,
                include=("rank", "ranking", "grossing", "brand"),
                url_include=("rank",),
                prefer_lower_values=True,
            )
        if not result.get("appstore_review_trend"):
            result["appstore_review_trend"] = _pick_series(
                series_candidates,
                include=("comment", "review", "rating"),
                url_include=("comment",),
            )
        if not result.get("dau_trend_90d"):
            result["dau_trend_90d"] = _pick_series(
                series_candidates,
                include=("dau", "active", "act", "\u65e5\u6d3b"),
                url_include=("appstatus", "status", "active"),
            )
        if not result.get("downloads_trend_90d"):
            result["downloads_trend_90d"] = _pick_series(
                series_candidates,
                include=("download", "downloads", "estimate"),
                url_include=("download", "pred"),
            )
        if not result.get("revenue_trend_90d"):
            result["revenue_trend_90d"] = _pick_series(
                series_candidates,
                include=("revenue", "income", "sales"),
                url_include=("revenue", "income", "pred"),
            )

        if not result.get("grossing_rank_cn"):
            rank = _find_scalar(scalar_candidates, ("grossing_rank", "grossing"))
            if rank is None:
                rank = _find_app_rank_in_payloads(self.api_payloads, self.app_id, result.get("app_name", ""))
            if rank is not None:
                result["grossing_rank_cn"] = rank
                result["grossing_rank"] = rank

        _extract_pred_estimate_scalars(self.api_payloads, result)

        result["api_urls"] = sorted({item["url"].split("?")[0] for item in self.api_payloads})
        if "api.qimai.cn" not in joined_by_url and not self.api_payloads:
            self.warnings.append("No api.qimai.cn JSON responses were captured.")

    def _extract_chart_fields(self, result: dict[str, Any]) -> None:
        if not self.chart_series:
            return
        result["qimai_chart_sources"] = {
            page_name: len(series)
            for page_name, series in sorted(self.chart_series.items())
            if series
        }
        api_urls = list(result.get("api_urls", [])) if isinstance(result.get("api_urls"), list) else []
        api_urls.extend(f"qimai_chart://{page_name}" for page_name, series in self.chart_series.items() if series)
        result["api_urls"] = sorted(set(api_urls))
        mapping = {
            "rank": "ios_grossing_rank_trend",
            "comment": "appstore_review_trend",
            "appstatus": "dau_trend_90d",
            "download": "downloads_trend_90d",
            "income": "revenue_trend_90d",
        }
        for page_name, target_key in mapping.items():
            if result.get(target_key):
                continue
            series = self.chart_series.get(page_name, [])
            if series:
                result[target_key] = series

    def _extract_export_fields(self, result: dict[str, Any]) -> None:
        if not self.export_rows:
            return
        result["qimai_export_sources"] = {
            page_name: len(rows)
            for page_name, rows in sorted(self.export_rows.items())
            if rows
        }
        api_urls = list(result.get("api_urls", [])) if isinstance(result.get("api_urls"), list) else []
        api_urls.extend(f"qimai_export://{page_name}" for page_name, rows in self.export_rows.items() if rows)
        result["api_urls"] = sorted(set(api_urls))
        mapping = {
            "rank": "ios_grossing_rank_trend",
            "comment": "appstore_review_trend",
            "download": "downloads_trend_90d",
            "income": "revenue_trend_90d",
        }
        for page_name, target_key in mapping.items():
            if result.get(target_key):
                continue
            series = _series_from_export_rows(self.export_rows.get(page_name, []), page_name)
            if series:
                result[target_key] = series

    def _finalize_derived_fields(self, result: dict[str, Any]) -> None:
        for key in ("ios_grossing_rank_trend", "appstore_review_trend", "dau_trend_90d", "downloads_trend_90d", "revenue_trend_90d"):
            if result.get(key):
                result[key] = _trim_series(_sort_series(_sanitize_qimai_metric_series(key, result[key])), 90)

        if not result.get("grossing_rank_cn") and result.get("ios_grossing_rank_trend"):
            latest_rank = _latest_series_value(result.get("ios_grossing_rank_trend", []))
            if latest_rank is not None:
                result["grossing_rank_cn"] = f"#{int(latest_rank)}"
                result["grossing_rank"] = result["grossing_rank_cn"]
                result["grossing_rank_source"] = "qimai_grossing_rank_trend_latest"

        result["dau_avg_30d"] = result.get("dau_avg_30d") or _average_last_n(result.get("dau_trend_90d", []), 30)
        result["downloads_avg_30d"] = result.get("downloads_avg_30d") or _average_last_n(result.get("downloads_trend_90d", []), 30)
        result["revenue_avg_30d"] = result.get("revenue_avg_30d") or _average_last_n(result.get("revenue_trend_90d", []), 30)


def _normalize_app_id(value: Any) -> str:
    text = str(value or "").strip()
    digits = re.sub(r"\D+", "", text)
    return digits or text


def _desktop_user_agent() -> str:
    return (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
    )


def _stealth_script() -> str:
    return """
        Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
        window.navigator.chrome = { runtime: {} };
    """


def _qimai_pred_params(app_id: str, country: str) -> dict[str, Any]:
    end_date = date.today() - timedelta(days=1)
    start_date = end_date - timedelta(days=89)
    return {
        "appid": app_id,
        "country": country,
        "platform": "iphone",
        "sdate": start_date.isoformat(),
        "edate": end_date.isoformat(),
        "_timeout": 60000,
    }


def _qimai_direct_metric_params(app_id: str, country: str) -> dict[str, Any]:
    end_date = date.today()
    start_date = end_date - timedelta(days=90)
    return {
        "appid": app_id,
        "country": country,
        "device": "iphone",
        "sdate": start_date.isoformat(),
        "edate": end_date.isoformat(),
        "_timeout": 60000,
    }


def _qimai_direct_metrics_script() -> str:
    return """
        async (params) => {
            const app = document.querySelector('#app');
            const vm = app && app.__vue__;
            const http = vm && (vm.$http || (vm.$root && vm.$root.$http));
            if (!http) {
                return { error: 'Qimai Vue $http is unavailable' };
            }
            const call = async (name, path, extra) => {
                const requestParams = Object.assign({}, params, extra || {});
                delete requestParams._timeout;
                try {
                    const response = await http.get(path, { params: requestParams, timeout: params._timeout || 60000 });
                    return {
                        url: 'https://api.qimai.cn' + path,
                        payload: response && response.data ? response.data : response,
                    };
                } catch (error) {
                    return {
                        url: 'https://api.qimai.cn' + path,
                        payload: error && error.response && error.response.data
                            ? error.response.data
                            : { code: 'client_error', msg: String(error) },
                    };
                }
            };
            return {
                rankMoreGrossing: await call('rankMoreGrossing', '/app/rankMore', {
                    export_type: 'app_rank',
                    brand: 'grossing',
                    day: 1,
                    appRankShow: 1,
                    subclass: 'all',
                    simple: 1,
                    rankEchartType: 1,
                }),
                rankGrossing: await call('rankGrossing', '/app/rank', {
                    export_type: 'app_rank',
                    brand: 'grossing',
                    day: 0,
                    appRankShow: 1,
                    subclass: 'all',
                    simple: 1,
                }),
                commentRateNum90d: await call('commentRateNum90d', '/app/commentRateNum', {
                    export_type: 'comment_rate_num',
                    typec: 'day',
                }),
                appStatus90d: await call('appStatus90d', '/app/appStatusList', {
                    export_type: 'app_status',
                    type: 'dau',
                    typec: 'day',
                }),
            };
        }
    """


def _qimai_pred_estimate_script() -> str:
    return """
        async (params) => {
            const app = document.querySelector('#app');
            const vm = app && app.__vue__;
            const http = vm && (vm.$http || (vm.$root && vm.$root.$http));
            if (!http) {
                return { error: 'Qimai Vue $http is unavailable' };
            }
            const call = async (name, path) => {
                try {
                    const response = await http.get(path, { params });
                    return {
                        url: 'https://api.qimai.cn' + path,
                        payload: response && response.data ? response.data : response,
                    };
                } catch (error) {
                    return {
                        url: 'https://api.qimai.cn' + path,
                        payload: error && error.response && error.response.data
                            ? error.response.data
                            : { code: 'client_error', msg: String(error) },
                    };
                }
            };
            return {
                download: await call('download', '/pred/download'),
                revenue: await call('revenue', '/pred/revenue'),
                income: await call('income', '/pred/income'),
            };
        }
    """


def _qimai_echarts_extract_script() -> str:
    return """
        (pageName) => {
            const echartsApi = window.echarts;
            const isTimestampLike = (num) => {
                const abs = Math.abs(Number(num));
                return (abs >= 1000000000 && abs <= 4102444800)
                    || (abs >= 1000000000000 && abs <= 4102444800000);
            };
            const valueOf = (item) => {
                if (item == null) return null;
                if (typeof item === 'number') return item;
                if (Array.isArray(item)) {
                    const nums = [];
                    for (let i = 0; i < item.length; i += 1) {
                        const num = Number(String(item[i]).replace(/,/g, ''));
                        if (Number.isFinite(num)) nums.push(num);
                    }
                    if (!nums.length) return null;
                    const nonTimestamp = nums.filter((num) => !isTimestampLike(num));
                    if (pageName === 'rank') {
                        return nonTimestamp.find((num) => num > 0 && num <= 2000) ?? null;
                    }
                    return nonTimestamp.length ? nonTimestamp[nonTimestamp.length - 1] : null;
                }
                if (typeof item === 'object') {
                    if ('value' in item) return valueOf(item.value);
                    for (const key of ['num', 'count', 'rank', 'income', 'revenue', 'download', 'dau', 'y']) {
                        if (key in item) return valueOf(item[key]);
                    }
                }
                const num = Number(String(item).replace(/,/g, ''));
                return Number.isFinite(num) ? num : null;
            };
            const textOf = (value) => String(value == null ? '' : value).toLowerCase();
            const chooseSeries = (series) => {
                const tokens = {
                    rank: ['畅销', 'grossing', 'rank'],
                    appstatus: ['dau', '日活', '活跃'],
                    download: ['下载', 'download'],
                    income: ['收入', 'revenue', 'income'],
                }[pageName] || [];
                for (const token of tokens) {
                    const found = series.find((item) => textOf(item.name).includes(textOf(token)));
                    if (found) return found;
                }
                if (pageName === 'rank') {
                    const ranked = series.find((item) => (item.data || []).some((point) => {
                        const num = valueOf(point);
                        return num != null && num > 0 && num <= 2000;
                    }));
                    if (ranked) return ranked;
                }
                return series.find((item) => Array.isArray(item.data) && item.data.length) || null;
            };
            const instances = [];
            const nodes = Array.from(document.querySelectorAll('div, canvas'));
            for (const node of nodes) {
                const key = Object.keys(node).find((name) => name.indexOf('_echarts_instance_') >= 0);
                const id = key ? node[key] : node.getAttribute && node.getAttribute('_echarts_instance_');
                if (!id || !echartsApi || !echartsApi.getInstanceById) continue;
                const instance = echartsApi.getInstanceById(id);
                if (instance && !instances.includes(instance)) instances.push(instance);
            }
            const charts = instances.map((instance) => instance.getOption && instance.getOption()).filter(Boolean);
            for (const option of charts) {
                const xAxis = Array.isArray(option.xAxis) ? option.xAxis[0] : option.xAxis;
                const dates = xAxis && Array.isArray(xAxis.data) ? xAxis.data : [];
                const series = Array.isArray(option.series) ? option.series : [];
                if (!dates.length || !series.length) continue;
                if (pageName === 'comment') {
                    return dates.map((date, index) => {
                        let total = 0;
                        let seen = false;
                        for (const item of series) {
                            const num = valueOf((item.data || [])[index]);
                            if (num != null) {
                                total += num;
                                seen = true;
                            }
                        }
                        return seen ? { date, value: total } : null;
                    }).filter(Boolean);
                }
                const selected = chooseSeries(series);
                if (!selected) continue;
                return dates.map((date, index) => {
                    const value = valueOf((selected.data || [])[index]);
                    return value == null ? null : { date, value };
                }).filter(Boolean);
            }
            return [];
        }
    """


def _append_pred_estimate_payloads(state: "_QimaiCaptureState", responses: Any) -> None:
    if not isinstance(responses, dict):
        state.warnings.append("Qimai pred estimate request returned an invalid response.")
        return
    if responses.get("error"):
        state.warnings.append(str(responses["error"]))
        return
    for key in ("download", "revenue", "income"):
        item = responses.get(key)
        if not isinstance(item, dict):
            continue
        url = str(item.get("url") or f"https://api.qimai.cn/pred/{key}")
        state._append_payload(url, item.get("payload"))


def _append_direct_metric_payloads(state: "_QimaiCaptureState", responses: Any) -> None:
    if not isinstance(responses, dict):
        state.warnings.append("Qimai direct metric request returned an invalid response.")
        return
    if responses.get("error"):
        state.warnings.append(str(responses["error"]))
        return
    fallback_urls = {
        "rankMoreGrossing": "https://api.qimai.cn/app/rankMore",
        "rankGrossing": "https://api.qimai.cn/app/rank",
        "commentRateNum90d": "https://api.qimai.cn/app/commentRateNum",
        "appStatus90d": "https://api.qimai.cn/app/appStatusList",
    }
    for key, fallback_url in fallback_urls.items():
        item = responses.get(key)
        if not isinstance(item, dict):
            continue
        url = str(item.get("url") or fallback_url)
        state._append_payload(url, item.get("payload"))


async def _click_text_async(page: Any, text: str, delay: float = 0.7, jitter: float = 0.0) -> None:
    try:
        await page.get_by_text(text, exact=True).first.click(timeout=1500)
        wait_seconds = _throttle_seconds(delay, min(jitter, delay))
        logger.debug("[Qimai] throttle after click '{}': {:.2f}s", text, wait_seconds)
        await asyncio.sleep(wait_seconds)
    except Exception:
        return


def _click_text_sync(page: Any, text: str, delay: float = 0.7, jitter: float = 0.0) -> None:
    try:
        page.get_by_text(text, exact=True).first.click(timeout=1500)
        wait_seconds = _throttle_seconds(delay, min(jitter, delay))
        logger.debug("[Qimai] throttle after click '{}': {:.2f}s", text, wait_seconds)
        time.sleep(wait_seconds)
    except Exception:
        return


async def _click_filter_option_async(page: Any, group_label: str, option_label: str, delay: float = 0.7, jitter: float = 0.0) -> bool:
    try:
        clicked = await page.evaluate(_qimai_click_filter_script(), {"group": group_label, "option": option_label})
        if clicked:
            wait_seconds = _throttle_seconds(delay, min(jitter, delay))
            logger.debug("[Qimai] throttle after filter click '{}:{}': {:.2f}s", group_label, option_label, wait_seconds)
            await asyncio.sleep(wait_seconds)
            return True
    except Exception as exc:
        logger.debug("[Qimai] filter click failed '{}:{}': {}", group_label, option_label, exc)
    await _click_text_async(page, option_label, delay, jitter)
    return False


def _click_filter_option_sync(page: Any, group_label: str, option_label: str, delay: float = 0.7, jitter: float = 0.0) -> bool:
    try:
        clicked = page.evaluate(_qimai_click_filter_script(), {"group": group_label, "option": option_label})
        if clicked:
            wait_seconds = _throttle_seconds(delay, min(jitter, delay))
            logger.debug("[Qimai] throttle after filter click '{}:{}': {:.2f}s", group_label, option_label, wait_seconds)
            time.sleep(wait_seconds)
            return True
    except Exception as exc:
        logger.debug("[Qimai] filter click failed '{}:{}': {}", group_label, option_label, exc)
    _click_text_sync(page, option_label, delay, jitter)
    return False


def _qimai_click_filter_script() -> str:
    return """
        ({ group, option }) => {
            const text = (node) => (node && node.innerText ? node.innerText : '').replace(/\\s+/g, ' ').trim();
            const visible = (node) => {
                if (!node) return false;
                const style = window.getComputedStyle(node);
                const rect = node.getBoundingClientRect();
                return style.visibility !== 'hidden' && style.display !== 'none' && rect.width > 0 && rect.height > 0;
            };
            const candidates = Array.from(document.querySelectorAll('section, .filter, .screen, .search-box, .rank-filter, .condition, .choose, .m-b, .item, .line, li, div'))
                .filter((node) => visible(node) && text(node).includes(group) && text(node).includes(option))
                .sort((a, b) => text(a).length - text(b).length);
            for (const container of candidates) {
                const nodes = Array.from(container.querySelectorAll('button, a, span, label, li, div'))
                    .filter((node) => visible(node) && text(node) === option)
                    .sort((a, b) => {
                        const activeA = /active|selected|cur|on/.test(String(a.className || '')) ? 1 : 0;
                        const activeB = /active|selected|cur|on/.test(String(b.className || '')) ? 1 : 0;
                        return activeA - activeB;
                    });
                if (nodes.length) {
                    nodes[0].click();
                    return true;
                }
            }
            const exact = Array.from(document.querySelectorAll('button, a, span, label, li, div'))
                .find((node) => visible(node) && text(node) === option);
            if (exact) {
                exact.click();
                return true;
            }
            return false;
        }
    """


async def _click_export_button_async(page: Any) -> bool:
    for locator in _export_button_locators(page):
        try:
            count = await locator.count()
        except Exception:
            continue
        for index in reversed(range(min(count, 6))):
            try:
                candidate = locator.nth(index)
                if not await candidate.is_visible(timeout=500):
                    continue
                await candidate.click(timeout=1500)
                return True
            except Exception:
                continue
    return False


def _click_export_button_sync(page: Any) -> bool:
    for locator in _export_button_locators(page):
        try:
            count = locator.count()
        except Exception:
            continue
        for index in reversed(range(min(count, 6))):
            try:
                candidate = locator.nth(index)
                if not candidate.is_visible(timeout=500):
                    continue
                candidate.click(timeout=1500)
                return True
            except Exception:
                continue
    return False


def _export_button_locators(page: Any) -> list[Any]:
    return [
        page.get_by_text("\u5bfc\u51fa\u6570\u636e", exact=True),
        page.get_by_text("\u5bfc\u51fa", exact=True),
        page.locator("[title*='\u5bfc\u51fa'], [aria-label*='\u5bfc\u51fa']"),
        page.locator(".icon-download, .el-icon-download, .download, .download-btn"),
        page.locator("[class*='download'], [class*='export']"),
        page.locator("button:has-text('\u5bfc\u51fa'), a:has-text('\u5bfc\u51fa')"),
    ]


async def _click_table_view_async(page: Any) -> None:
    for locator in _table_view_locators(page):
        try:
            candidate = locator.first
            if await candidate.is_visible(timeout=500):
                await candidate.click(timeout=1000)
                await asyncio.sleep(0.5)
                return
        except Exception:
            continue


def _click_table_view_sync(page: Any) -> None:
    for locator in _table_view_locators(page):
        try:
            candidate = locator.first
            if candidate.is_visible(timeout=500):
                candidate.click(timeout=1000)
                time.sleep(0.5)
                return
        except Exception:
            continue


def _table_view_locators(page: Any) -> list[Any]:
    return [
        page.get_by_text("\u8868\u683c", exact=True),
        page.get_by_text("\u5217\u8868", exact=True),
        page.locator("[title*='\u8868\u683c'], [aria-label*='\u8868\u683c'], [title*='\u5217\u8868'], [aria-label*='\u5217\u8868']"),
        page.locator(".icon-table, .icon-list, .el-icon-s-grid, .el-icon-s-unfold"),
    ]


def _throttle_seconds(base_delay: float, jitter: float) -> float:
    base = max(float(base_delay or 0), 0.0)
    spread = max(float(jitter or 0), 0.0)
    return base + random.uniform(0, spread)


async def _safe_scroll_async(page: Any) -> None:
    try:
        await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
    except Exception:
        return


def _safe_scroll_sync(page: Any) -> None:
    try:
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
    except Exception:
        return


def _first_nonempty_line(text: str) -> str:
    ignored = {"\u4e03\u9ea6\u6570\u636e", "qimai", "\u767b\u5f55", "\u6ce8\u518c"}
    for raw in (text or "").splitlines():
        line = raw.strip()
        if not line or line.lower() in ignored:
            continue
        if len(line) > 40:
            continue
        return line
    return ""


def _looks_like_navigation_text(value: Any) -> bool:
    text = str(value or "").strip()
    return text in {
        "\u4e03\u9ea6\u6570\u636e",
        "\u699c\u5355",
        "\u5de5\u5177",
        "Apple Ads",
        "\u7814\u7a76\u9662",
        "\u4e03\u9ea6\u670d\u52a1",
        "NextWorld",
        "\u767b\u5f55",
        "\u6ce8\u518c",
    }


def _extract_rating(text: str) -> tuple[Any | None, int | None]:
    rating = None
    rating_count = None
    count_match = re.search(r"([0-9,.\u4e07]+)\s*\u4e2a\u8bc4\u5206", text)
    if count_match:
        rating_count = _to_int(count_match.group(1))
        after_count = text[count_match.end() : count_match.end() + 80]
        rating_match = re.search(r"([0-5](?:\.\d+)?)", after_count)
        if rating_match:
            rating = _to_number(rating_match.group(1))
    if rating is None:
        rating_match = re.search(r"\u8bc4\u5206\s*([0-5](?:\.\d+)?)", text)
        if rating_match:
            rating = _to_number(rating_match.group(1))
    return rating, rating_count


def _extract_rank_near_label(text: str, label: str) -> str:
    compact = re.sub(r"\s+", " ", text or "")
    patterns = [
        rf"{re.escape(label)}[^0-9\u7b2c-]{{0,20}}\u7b2c\s*([0-9,]+)\s*\u540d",
        rf"\u7b2c\s*([0-9,]+)\s*\u540d[^。；;\n]{{0,20}}{re.escape(label)}",
    ]
    for pattern in patterns:
        match = re.search(pattern, compact)
        if match:
            return f"#{match.group(1).replace(',', '')}"
    return ""


def _extract_grossing_rank_from_public_rank(text: str, app_name: str) -> str:
    if not text or not app_name:
        return ""
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if not lines:
        return ""

    section_start = 0
    for index, line in enumerate(lines):
        if "\u7545\u9500\u699c" in line:
            section_start = index + 1
            break
    section = lines[section_start:]
    for index, line in enumerate(section):
        if app_name not in line:
            continue
        for previous in reversed(section[max(0, index - 4):index]):
            if re.fullmatch(r"\d{1,3}", previous):
                return f"#{previous}"
        match = re.search(rf"(?:^|\s)(\d{{1,3}})\s+{re.escape(app_name)}", " ".join(section[max(0, index - 4):index + 1]))
        if match:
            return f"#{match.group(1)}"
    return ""


def _extract_yesterday_downloads(text: str) -> int | None:
    match = re.search(r"\u6628\u65e5\u4e0b\u8f7d\u91cf\s*([0-9,.\u4e07]+)", text or "")
    if not match:
        return None
    return _to_int(match.group(1))


def _walk_json(value: Any, path: str = "") -> Iterable[tuple[str, Any]]:
    if isinstance(value, dict):
        for key, child in value.items():
            child_path = f"{path}.{key}" if path else str(key)
            yield child_path, child
            yield from _walk_json(child, child_path)
    elif isinstance(value, list):
        for index, child in enumerate(value):
            child_path = f"{path}[{index}]"
            yield child_path, child
            yield from _walk_json(child, child_path)


def _find_scalar(candidates: Iterable[tuple[str, Any]], keys: tuple[str, ...], *, prefer_float: bool = False) -> Any | None:
    for path, value in candidates:
        lowered = path.lower()
        if not any(key in lowered for key in keys):
            continue
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return float(value) if prefer_float else value
        if isinstance(value, str):
            converted = _to_number(value)
            if converted is not None:
                return converted
    return None


def _find_app_rank_in_payloads(api_payloads: list[dict[str, Any]], app_id: str, app_name: str = "") -> int | None:
    for item in api_payloads:
        if _qimai_api_category(str(item.get("url", ""))) != "rank":
            continue
        rank = _find_app_rank_in_node(item.get("payload"), app_id, app_name)
        if rank is not None:
            return rank
    return None


def _find_app_rank_in_node(node: Any, app_id: str, app_name: str = "") -> int | None:
    if isinstance(node, list):
        for child in node:
            rank = _find_app_rank_in_node(child, app_id, app_name)
            if rank is not None:
                return rank
        return None
    if not isinstance(node, dict):
        return None

    if _dict_matches_qimai_app(node, app_id, app_name):
        rank = _rank_value_from_dict(node)
        if rank is not None:
            return rank

    for child in node.values():
        rank = _find_app_rank_in_node(child, app_id, app_name)
        if rank is not None:
            return rank
    return None


def _dict_matches_qimai_app(item: dict[str, Any], app_id: str, app_name: str = "") -> bool:
    target_id = str(app_id or "").strip()
    if target_id:
        for key, value in item.items():
            lowered = str(key).lower()
            if lowered in {"appid", "app_id", "appstore_id", "apple_id", "trackid", "track_id", "id"} and _normalize_app_id(value) == target_id:
                return True
    target_name = str(app_name or "").strip()
    if target_name and not _looks_like_navigation_text(target_name):
        for key, value in item.items():
            lowered = str(key).lower()
            if lowered in {"name", "appname", "app_name", "title"} and target_name in str(value):
                return True
    return False


def _rank_value_from_dict(item: dict[str, Any]) -> int | None:
    rank_keys = (
        "rank",
        "ranking",
        "rank_num",
        "ranknum",
        "current_rank",
        "currentrank",
        "sort",
        "no",
    )
    for key in rank_keys:
        for raw_key, value in item.items():
            if str(raw_key).lower() != key:
                continue
            rank = _to_int(value)
            if rank is not None and 0 < rank <= 2000:
                return rank
    return None


def _extract_series_candidates(api_payloads: list[dict[str, Any]]) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    for item in api_payloads:
        url = item.get("url", "")
        api_category = _qimai_api_category(url)
        for path, value in _walk_json(item.get("payload")):
            if not isinstance(value, list) or len(value) < 2:
                continue
            series = _normalize_series(value)
            if len(series) >= 2:
                candidates.append({"url": url, "path": path, "category": api_category, "series": series})
    return candidates


def _extract_pred_estimate_scalars(api_payloads: list[dict[str, Any]], result: dict[str, Any]) -> None:
    for item in api_payloads:
        category = _qimai_api_category(str(item.get("url", "")))
        if category not in {"download", "revenue"}:
            continue
        payload = item.get("payload")
        data = payload.get("data") if isinstance(payload, dict) else None
        if not isinstance(data, dict):
            continue
        if category == "download":
            result["downloads_avg_30d"] = result.get("downloads_avg_30d") or _first_metric(
                data,
                (
                    "monthDownloadIphoneAvg",
                    "monthDownloadAvg",
                    "monthDownloadIpadAvg",
                    "month_download_iphone_avg",
                    "month_download_avg",
                ),
            )
            result["yesterday_downloads"] = result.get("yesterday_downloads") or _first_metric(
                data,
                (
                    "yesterdayDownloadIphone",
                    "yesterdayDownload",
                    "yesterdayDownloadIpad",
                    "yesterday_download_iphone",
                    "yesterday_download",
                ),
            )
        elif category == "revenue":
            result["revenue_avg_30d"] = result.get("revenue_avg_30d") or _first_metric(
                data,
                (
                    "monthRevenueIphoneAvg",
                    "monthRevenueAvg",
                    "monthRevenueIpadAvg",
                    "month_revenue_iphone_avg",
                    "month_revenue_avg",
                ),
            )


def _first_metric(data: dict[str, Any], keys: tuple[str, ...]) -> int | float | None:
    for key in keys:
        if key not in data:
            continue
        value = _to_number(data.get(key))
        if value is not None:
            return value
    return None


def _read_qimai_export_rows(path_value: Any, suggested_filename: str = "") -> list[dict[str, Any]]:
    if not path_value:
        return []
    path = Path(str(path_value))
    suffix = (path.suffix or Path(str(suggested_filename or "")).suffix).lower()
    try:
        if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            return _read_qimai_xlsx_rows(path)
        return _read_qimai_csv_rows(path)
    except Exception as exc:
        logger.debug("[Qimai] failed to parse export {}: {}", suggested_filename or path, exc)
        return []


def _read_qimai_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.debug("[Qimai] openpyxl is unavailable; cannot parse xlsx export")
        return []
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        raw_rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    return _table_rows_from_values(raw_rows)


def _read_qimai_csv_rows(path: Path) -> list[dict[str, Any]]:
    for encoding in ("utf-8-sig", "gb18030", "utf-16"):
        try:
            with open(path, "r", encoding=encoding, newline="") as handle:
                sample = handle.read(2048)
                handle.seek(0)
                dialect = csv.Sniffer().sniff(sample) if sample else csv.excel
                return _table_rows_from_values(list(csv.reader(handle, dialect)))
        except Exception:
            continue
    return []


def _table_rows_from_values(raw_rows: list[Any]) -> list[dict[str, Any]]:
    rows = [
        ["" if cell is None else cell for cell in row]
        for row in raw_rows
        if row and any(str(cell or "").strip() for cell in row)
    ]
    if not rows:
        return []

    header_index = 0
    for index, row in enumerate(rows[:8]):
        if any(_looks_like_date_header(cell) for cell in row):
            header_index = index
            break
        if any(_normalize_date(cell) for cell in row):
            header_index = max(index - 1, 0)
            break

    headers = [str(cell or "").strip() or f"column_{idx}" for idx, cell in enumerate(rows[header_index])]
    normalized: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        item = {
            headers[index] if index < len(headers) else f"column_{index}": cell
            for index, cell in enumerate(row)
        }
        if item:
            normalized.append(item)
    return normalized


def _looks_like_date_header(value: Any) -> bool:
    lowered = str(value or "").strip().lower()
    return lowered in {"date", "day", "time", "\u65e5\u671f", "\u65f6\u95f4"} or "\u65e5\u671f" in lowered


def _series_from_export_rows(rows: list[dict[str, Any]], page_name: str) -> list[dict[str, Any]]:
    series: list[dict[str, Any]] = []
    for row in rows:
        point = _point_from_export_row(row, page_name)
        if point:
            series.append(point)
    return _dedupe_series(series)


def _point_from_export_row(row: dict[str, Any], page_name: str) -> dict[str, Any] | None:
    date_text = ""
    for key, value in row.items():
        if _looks_like_date_header(key) or _normalize_date(value):
            date_text = _normalize_date(value)
            if date_text:
                break
    if not date_text:
        return None

    value = _export_metric_value(row, page_name)
    if value is None:
        return None
    return {"date": date_text, "value": value}


def _export_metric_value(row: dict[str, Any], page_name: str) -> int | float | None:
    preferences = {
        "rank": ("rank", "ranking", "\u6392\u540d", "\u7545\u9500"),
        "download": ("download", "downloads", "\u4e0b\u8f7d"),
        "income": ("income", "revenue", "sales", "\u6536\u5165", "\u9500\u552e"),
        "comment": ("\u5408\u8ba1", "total", "comment", "review", "\u8bc4\u8bba", "\u8bc4\u5206"),
    }
    for key, value in row.items():
        key_lower = str(key or "").lower()
        if _looks_like_date_header(key_lower):
            continue
        if any(token.lower() in key_lower for token in preferences.get(page_name, ())):
            number = _to_number(value)
            if number is not None:
                return number

    numeric_values: list[int | float] = []
    for key, value in row.items():
        if _looks_like_date_header(key):
            continue
        number = _to_number(value)
        if number is not None:
            numeric_values.append(number)
    if not numeric_values:
        return None
    if page_name == "comment":
        return sum(numeric_values)
    return numeric_values[0]


def _qimai_api_category(url: str) -> str:
    lowered = (url or "").lower()
    category_tokens = {
        "rank": ("rank", "ranking", "brand"),
        "comment": ("comment", "review"),
        "appstatus": ("appstatus", "active", "dau", "status"),
        "download": ("download", "downloads", "down"),
        "revenue": ("revenue", "income", "sales"),
    }
    for category, tokens in category_tokens.items():
        if any(token in lowered for token in tokens):
            return category
    return ""


def _normalize_series(items: list[Any]) -> list[dict[str, Any]]:
    series: list[dict[str, Any]] = []
    for item in items:
        point = _normalize_point(item)
        if point:
            series.append(point)
    return _dedupe_series(series)


def _normalize_point(item: Any) -> dict[str, Any] | None:
    if isinstance(item, dict):
        date_value = _first_key_value(item, ("date", "day", "time", "timestamp", "dt", "x"))
        metric_value = _first_key_value(
            item,
            ("value", "num", "count", "rank", "ranking", "income", "revenue", "download", "downloads", "dau", "y"),
        )
        value = _to_number(metric_value)
        if metric_value is None or (value is not None and _looks_like_timestamp_number(value)):
            fallback_value = None
            for key, value in item.items():
                if value == date_value:
                    continue
                fallback_number = _first_non_timestamp_number(value)
                if fallback_number is not None:
                    fallback_value = fallback_number
                    break
            if fallback_value is not None:
                metric_value = fallback_value
        date_text = _normalize_date(date_value)
        value = _to_number(metric_value)
        if date_text and value is not None:
            return {"date": date_text, "value": value}
    elif isinstance(item, (list, tuple)) and len(item) >= 2:
        date_text = ""
        date_index = -1
        for index, raw_value in enumerate(item):
            candidate_date = _normalize_date(raw_value)
            if candidate_date:
                date_text = candidate_date
                date_index = index
                break
        value = None
        for index, raw_value in enumerate(item):
            if index == date_index:
                continue
            candidate_value = _to_number(raw_value)
            if candidate_value is not None and not _looks_like_timestamp_number(candidate_value):
                value = candidate_value
                break
        if date_text and value is not None:
            return {"date": date_text, "value": value}
    return None


def _first_key_value(item: dict[str, Any], keys: tuple[str, ...]) -> Any:
    lowered = {str(key).lower(): value for key, value in item.items()}
    for key in keys:
        if key in lowered:
            return lowered[key]
    for raw_key, value in item.items():
        raw_lower = str(raw_key).lower()
        if any(key in raw_lower for key in keys):
            return value
    return None


def _normalize_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        timestamp = float(value)
        if not _looks_like_timestamp_number(timestamp):
            return ""
        if timestamp > 10_000_000_000:
            timestamp = timestamp / 1000
        try:
            return datetime.fromtimestamp(timestamp, tz=timezone.utc).date().isoformat()
        except Exception:
            return ""
    text = str(value or "").strip()
    if not text:
        return ""
    match = re.search(r"20\d{2}[-/.]\d{1,2}[-/.]\d{1,2}", text)
    if match:
        return match.group(0).replace("/", "-").replace(".", "-")
    match = re.search(r"(\d{1,2})[-/.](\d{1,2})", text)
    if match:
        current_year = date.today().year
        return f"{current_year}-{int(match.group(1)):02d}-{int(match.group(2)):02d}"
    match = re.search(r"(\d{1,2})\s*\u6708\s*(\d{1,2})\s*\u65e5?", text)
    if match:
        current_year = date.today().year
        return f"{current_year}-{int(match.group(1)):02d}-{int(match.group(2)):02d}"
    return ""


def _looks_like_timestamp_number(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    try:
        number = abs(float(value))
    except (TypeError, ValueError):
        return False
    return (1_000_000_000 <= number <= 4_102_444_800) or (1_000_000_000_000 <= number <= 4_102_444_800_000)


def _first_non_timestamp_number(value: Any) -> int | float | None:
    if isinstance(value, dict):
        for key in ("rank", "ranking", "value", "num", "count", "income", "revenue", "download", "downloads", "dau", "y"):
            if key in value:
                number = _first_non_timestamp_number(value.get(key))
                if number is not None:
                    return number
        for child in value.values():
            number = _first_non_timestamp_number(child)
            if number is not None:
                return number
        return None
    if isinstance(value, (list, tuple)):
        for child in value:
            number = _first_non_timestamp_number(child)
            if number is not None:
                return number
        return None
    number = _to_number(value)
    if number is None or _looks_like_timestamp_number(number):
        return None
    return number


def _sanitize_qimai_metric_series(metric_key: str, series: list[dict[str, Any]]) -> list[dict[str, Any]]:
    sanitized: list[dict[str, Any]] = []
    for point in series:
        if not isinstance(point, dict):
            continue
        value = _to_number(point.get("value"))
        if value is None or _looks_like_timestamp_number(value):
            continue
        if metric_key == "ios_grossing_rank_trend" and not (0 < float(value) <= 2000):
            continue
        sanitized.append({"date": point.get("date", ""), "value": value})
    return sanitized


def _to_number(value: Any) -> float | int | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip().replace(",", "")
    if not text or text in {"-", "--"}:
        return None
    multiplier = 1
    if "\u4e07" in text:
        multiplier = 10000
        text = text.replace("\u4e07", "")
    if "\u4ebf" in text:
        multiplier = 100000000
        text = text.replace("\u4ebf", "")
    text = re.sub(r"[^0-9.\-]", "", text)
    if not text or text in {"-", "."}:
        return None
    try:
        number = float(text) * multiplier
    except ValueError:
        return None
    return int(number) if number.is_integer() else number


def _to_int(value: Any) -> int | None:
    number = _to_number(value)
    if number is None:
        return None
    return int(number)


def _pick_series(
    candidates: list[dict[str, Any]],
    *,
    include: tuple[str, ...],
    url_include: tuple[str, ...] = (),
    prefer_lower_values: bool = False,
) -> list[dict[str, Any]]:
    scored: list[tuple[int, dict[str, Any]]] = []
    expected_categories = _expected_qimai_categories(url_include)
    for candidate in candidates:
        category = str(candidate.get("category", ""))
        if expected_categories and category not in expected_categories:
            continue

        haystack = f"{candidate.get('path', '')}".lower()
        score = 0
        score += sum(5 for token in include if token.lower() in haystack)
        if category in expected_categories:
            score += 20
        score += min(len(candidate.get("series", [])), 90) // 10
        if prefer_lower_values:
            values = [point["value"] for point in candidate.get("series", []) if isinstance(point.get("value"), (int, float))]
            if values and max(values) <= 2000:
                score += 5
        if score > 0:
            scored.append((score, candidate))
    if not scored:
        return []
    scored.sort(key=lambda item: item[0], reverse=True)
    return scored[0][1]["series"]


def _expected_qimai_categories(url_include: tuple[str, ...]) -> set[str]:
    categories: set[str] = set()
    for token in url_include:
        category = _qimai_api_category(token)
        if category:
            categories.add(category)
    return categories


def _dedupe_series(series: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_date: dict[str, dict[str, Any]] = {}
    for point in series:
        by_date[point["date"]] = point
    return list(by_date.values())


def _sort_series(series: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(series, key=lambda point: point.get("date", ""))


def _trim_series(series: list[dict[str, Any]], days: int) -> list[dict[str, Any]]:
    cutoff = date.today() - timedelta(days=days + 7)
    trimmed: list[dict[str, Any]] = []
    for point in series:
        try:
            point_date = datetime.fromisoformat(str(point["date"])).date()
        except Exception:
            trimmed.append(point)
            continue
        if point_date >= cutoff:
            trimmed.append(point)
    return trimmed[-days:]


def _average_last_n(series: list[dict[str, Any]], n: int) -> float | None:
    values = [
        float(point["value"])
        for point in _sort_series(series)[-n:]
        if isinstance(point.get("value"), (int, float)) and not isinstance(point.get("value"), bool)
    ]
    if not values:
        return None
    return round(statistics.fmean(values), 2)


def _latest_series_value(series: list[dict[str, Any]]) -> int | float | None:
    for point in reversed(_sort_series(series)):
        if not isinstance(point, dict):
            continue
        value = point.get("value")
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return value
    return None


def _is_ignorable_playwright_capture_error(exc: BaseException) -> bool:
    if isinstance(exc, asyncio.CancelledError):
        return True
    name = exc.__class__.__name__
    text = str(exc).lower()
    return (
        name in {"TargetClosedError", "Error"}
        and (
            "target page, context or browser has been closed" in text
            or "browser has been closed" in text
            or "context has been closed" in text
            or "page has been closed" in text
        )
    )


def _series_debug_summary(candidates: list[dict[str, Any]]) -> list[dict[str, Any]]:
    summary = []
    for candidate in candidates[:20]:
        summary.append(
            {
                "url": candidate.get("url", "").split("?")[0],
                "path": candidate.get("path"),
                "category": candidate.get("category", ""),
                "points": len(candidate.get("series", [])),
            }
        )
    return summary
