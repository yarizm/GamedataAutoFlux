from datetime import datetime
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from src.storage.base import StorageRecord
from src.storage.factory import get_storage
from src.web.app import app


def test_youtube_export_matches_pipeline_record_keys(tmp_path, monkeypatch) -> None:
    task_id = "8a49f1c40413"
    record_key = f"{task_id}:https://www.youtube.com/@doudoureviews/videos:3"
    _save_record(
        StorageRecord(
            key=record_key,
            source="https://www.youtube.com/@doudoureviews/videos",
            data={
                "collector": "youtube_profiles",
                "input_url": "https://www.youtube.com/@doudoureviews/videos",
                "author_name": "Doudou Reviews",
                "channel_id": "UC-example",
                "channel_url": "https://www.youtube.com/channel/UC-example",
                "subscriber_count": "12345",
                "description": "reviews",
                "resolution_method": "handle",
                "resolution_status": "success",
            },
            metadata={
                "collector": "youtube_profiles",
                "task_id": task_id,
                "source_task": {
                    "task_id": task_id,
                    "collector_name": "youtube_profiles",
                },
            },
            stored_at=datetime(2026, 1, 1, 12, 0, 0),
        )
    )
    _save_record(
        StorageRecord(
            key="other-task:https://www.youtube.com/@other/videos:0",
            source="https://www.youtube.com/@other/videos",
            data={"collector": "youtube_profiles", "input_url": "https://www.youtube.com/@other/videos"},
            metadata={"collector": "youtube_profiles", "task_id": "other-task"},
            stored_at=datetime(2026, 1, 1, 12, 1, 0),
        )
    )

    monkeypatch.setattr("src.web.routes.youtube_export.EXPORT_DIR", tmp_path)

    with TestClient(app) as client:
        response = client.post(
            "/api/data/export/youtube",
            json={
                "collector": "youtube_profiles",
                "task_ids": [task_id],
                "format": "xlsx",
            },
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["record_count"] == 1
    filename = payload["download_url"].rsplit("/", 1)[-1]
    workbook_path = tmp_path / filename
    assert workbook_path.exists()

    workbook = load_workbook(workbook_path)
    sheet = workbook["博主信息"]
    assert sheet.max_row == 2
    assert sheet.cell(row=2, column=1).value == "https://www.youtube.com/@doudoureviews/videos"
    assert sheet.cell(row=2, column=2).value == "Doudou Reviews"


def test_youtube_export_can_fallback_to_task_id_from_record_key(tmp_path, monkeypatch) -> None:
    task_id = "legacy-task"
    _save_record(
        StorageRecord(
            key=f"{task_id}:https://www.youtube.com/watch?v=abc123DEF45:0",
            source="https://www.youtube.com/watch?v=abc123DEF45",
            data={
                "collector": "youtube_comments",
                "video_url": "https://www.youtube.com/watch?v=abc123DEF45",
                "title": "Video",
                "comments": [{"like_count": 7, "text": "hello", "published_at": "2026-01-01"}],
            },
            metadata={"collector": "youtube_comments"},
            stored_at=datetime(2026, 1, 1, 12, 0, 0),
        )
    )

    monkeypatch.setattr("src.web.routes.youtube_export.EXPORT_DIR", tmp_path)

    with TestClient(app) as client:
        response = client.post(
            "/api/data/export/youtube",
            json={
                "collector": "youtube_comments",
                "task_ids": [task_id],
                "format": "xlsx",
            },
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["record_count"] == 1
    workbook_path = tmp_path / Path(payload["download_url"]).name
    workbook = load_workbook(workbook_path)
    assert workbook["视频信息"].cell(row=2, column=1).value == "https://www.youtube.com/watch?v=abc123DEF45"
    assert workbook["评论信息"].cell(row=2, column=3).value == "hello"


def _save_record(record: StorageRecord) -> None:
    import asyncio

    async def run() -> None:
        store = get_storage()
        await store.initialize()
        try:
            await store.save(record)
        finally:
            await store.close()

    asyncio.run(run())
